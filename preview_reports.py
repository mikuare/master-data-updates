import base64
from datetime import datetime
from email.parser import BytesParser
from email.policy import default
from html import escape
from http.server import BaseHTTPRequestHandler, HTTPServer
from io import BytesIO
import json
import mimetypes
import os
from pathlib import Path
import shutil
import subprocess
import tempfile
from urllib.parse import parse_qs, quote, urlparse
from zoneinfo import ZoneInfo

import pandas as pd

from compare_employees import (
    DISPLAY_COLUMN_RENAMES,
    REPORT_FILENAMES,
    find_input_file,
    normalize_columns,
    process_reports,
    read_excel_file,
)


BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "web_uploads"
REPORT_FILES = list(REPORT_FILENAMES.values())
LAST_RUN_FILE = BASE_DIR / ".last_preview_run.txt"
PRINT_REPORT_QUEUE_FILE = BASE_DIR / ".print_report_queue.json"
PRINT_STATUS_BREAKDOWN_FILE = BASE_DIR / ".print_status_breakdown.json"
PRINT_PRESETS_FILE = BASE_DIR / ".print_report_presets.json"
PRINT_LOGO_BASE = BASE_DIR / "print_report_logo"
PRINT_LOGO_EXTENSIONS = (".png", ".jpg", ".jpeg", ".svg", ".webp")
PRINT_REPORT_ARCHIVE_DIR = BASE_DIR / "employee update print report"
UPDATED_SYSTEM_FILENAMES = ("updated_system_clean.xlsx", "updated_system_clean.xls")
UPDATED_SYSTEM_CONFIG_FILE = ".updated_system_config.json"
SYSTEM_BRANCH_GROUPS = {
    "QM BUILDERS": [
        "qm builders",
        "qm realty",
        "qmb production",
        "qmb hardware",
        "qmb equipment",
        "qmb construction",
        "qmb constructions",
    ],
    "ADAMANT": [
        "adamant",
        "adc construction",
        "adc constructions",
    ],
    "QM FARMS": [
        "qm farms",
        "qmb farm",
        "qmb farms",
    ],
    "QM DIVING RESORT": [
        "qm diving resort",
        "cafe de casilda",
        "diving resort",
    ],
    "QGDC": [
        "qgdc",
        "qgdc construction",
        "qgdc constructions",
    ],
    "QMAZ HOLDINGS": [
        "qmaz holdings",
        "qmaz operations",
    ],
}
UNMAPPED_BRANCH_LABEL = "UNMAPPED / OTHER"
UPDATED_SYSTEM_STATUS_COLUMN_ALIASES = {
    "status",
    "systemstatus",
    "employmentstatus",
    "employee_status",
    "employeestatus",
}
UPDATED_SYSTEM_BRANCH_COLUMN_ALIASES = {
    "branch",
    "branchbasis",
    "branchgroup",
    "grouping",
    "parentbranch",
    "company",
    "division",
    "group",
}
APP_TABS = {
    "compare-employees": "Compare Employees",
    "trace-duplicate-stock-items": "Trace Duplicate Stock Items",
}
REPORT_LABELS = {
    "inactive_to_update.xlsx": "Inactive To Update",
    "active_to_update.xlsx": "Active To Update",
    "new_active_employees.xlsx": "New Active Employees",
    "new_in_system_since_last_month.xlsx": "New In System Since Last Snapshot",
}
APP_TIMEZONE = ZoneInfo("Asia/Manila")


def local_now() -> datetime:
    return datetime.now(APP_TIMEZONE)


def format_local_timestamp(value: float | datetime, pattern: str = "%B %d, %Y %I:%M %p") -> str:
    if isinstance(value, datetime):
        dt = value
    else:
        dt = datetime.fromtimestamp(value, APP_TIMEZONE)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=APP_TIMEZONE)
    else:
        dt = dt.astimezone(APP_TIMEZONE)
    return dt.strftime(pattern)


def get_active_report_dir() -> Path:
    if LAST_RUN_FILE.exists():
        saved = LAST_RUN_FILE.read_text(encoding="utf-8").strip()
        if saved:
            path = Path(saved)
            if path.exists():
                return path
    return BASE_DIR


def set_active_report_dir(path: Path) -> None:
    LAST_RUN_FILE.write_text(str(path), encoding="utf-8")


def list_saved_report_runs() -> list[dict[str, str]]:
    if not UPLOAD_DIR.exists():
        return []

    runs: list[dict[str, str]] = []
    for path in sorted(
        (item for item in UPLOAD_DIR.iterdir() if item.is_dir() and item.name.startswith("run_")),
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    ):
        branch = get_hr_branch_label(path) or "Unknown branch"
        timestamp = format_local_timestamp(path.stat().st_mtime)
        runs.append(
            {
                "name": path.name,
                "branch": branch,
                "updated_at": timestamp,
            }
        )
    return runs


def get_saved_report_run(name: str) -> Path | None:
    candidate = (UPLOAD_DIR / Path(name).name).resolve()
    try:
        candidate.relative_to(UPLOAD_DIR.resolve())
    except ValueError:
        return None
    if candidate.exists() and candidate.is_dir() and candidate.name.startswith("run_"):
        return candidate
    return None


def load_print_report_queue() -> list[dict[str, object]]:
    if not PRINT_REPORT_QUEUE_FILE.exists():
        return []
    try:
        data = json.loads(PRINT_REPORT_QUEUE_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return []
    if not isinstance(data, list):
        return []
    rows = [item for item in data if isinstance(item, dict)]
    normalized = normalize_print_report_queue(rows)
    if normalized != rows:
        save_print_report_queue(normalized)
    return normalized


def save_print_report_queue(rows: list[dict[str, object]]) -> None:
    normalized = normalize_print_report_queue(rows)
    PRINT_REPORT_QUEUE_FILE.write_text(
        json.dumps(normalized, indent=2),
        encoding="utf-8",
    )


def load_print_status_breakdown() -> list[dict[str, object]]:
    if not PRINT_STATUS_BREAKDOWN_FILE.exists():
        return []
    try:
        data = json.loads(PRINT_STATUS_BREAKDOWN_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return []
    if not isinstance(data, list):
        return []
    return [item for item in data if isinstance(item, dict)]


def save_print_status_breakdown(rows: list[dict[str, object]]) -> None:
    PRINT_STATUS_BREAKDOWN_FILE.write_text(
        json.dumps(rows, indent=2),
        encoding="utf-8",
    )


def save_current_print_report_archive() -> Path:
    PRINT_REPORT_ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = format_local_timestamp(local_now(), "%Y%m%d_%H%M%S")
    target = PRINT_REPORT_ARCHIVE_DIR / f"employee_update_print_report_{timestamp}.pdf"
    target.write_bytes(export_print_report_pdf())
    return target


def default_print_preset() -> dict[str, str]:
    return {
        "preset_name": "Default",
        "company_name": "QM BUILDERS",
        "department_label": "QHSE",
        "form_title": "ACUMATICA EMPLOYEE UPDATE REPORT",
        "date_label": "DATE UPDATED EMPLOYEE MASTER DATA",
        "date_value": "",
        "section1_title": "Employee Update Totals",
        "section2_title": "Updated System Employee Status Totals",
        "section1_note": "Section 1 shows the employee update counts added from the comparison reports for each branch, including employee IDs newly added in the system since the last saved branch snapshot.",
        "section2_note": "Section 2 shows branch-based active, inactive, and overall employee totals from the uploaded updated system_clean file using the defined branch mapping.",
        "header_bg": "#e5f3f1",
        "status_header_bg": "#e8eefc",
        "branch_bg": "#fff7ed",
        "total_bg": "#fef3c7",
        "border_color": "#1f2937",
        "accent_color": "#0f766e",
        "page_padding_px": "24",
        "print_page_margin_mm": "10",
        "section_padding_px": "18",
        "table_cell_padding_px": "10",
        "section_gap_px": "18",
        "signature_top_mm": "25.4",
        "prepared_by_label": "Prepared by:",
        "prepared_by_name": "MR.JUDE MICHAEL MARTINEZ",
        "prepared_by_title": "database analyst",
        "verified_by_label": "Verified By:",
        "verified_by_name": "MR.PATRICK CASTILLO EROJO",
        "verified_by_title": "DIGITAL SYSTEM ANALYST",
    }


def load_print_presets() -> dict[str, object]:
    default = default_print_preset()
    if not PRINT_PRESETS_FILE.exists():
        return {
            "active": default["preset_name"],
            "presets": {
                default["preset_name"]: default,
            },
        }
    try:
        data = json.loads(PRINT_PRESETS_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {
            "active": default["preset_name"],
            "presets": {
                default["preset_name"]: default,
            },
        }
    presets = data.get("presets", {}) if isinstance(data, dict) else {}
    active = str(data.get("active", default["preset_name"])) if isinstance(data, dict) else default["preset_name"]
    if not isinstance(presets, dict):
        presets = {}
    normalized_presets: dict[str, dict[str, str]] = {default["preset_name"]: default}
    for name, preset in presets.items():
        if not isinstance(preset, dict):
            continue
        merged = default.copy()
        merged.update({key: str(value) for key, value in preset.items()})
        if "page_padding_px" not in preset and "page_padding_mm" in preset:
            merged["page_padding_px"] = str(preset["page_padding_mm"])
        if "section_padding_px" not in preset and "section_padding_mm" in preset:
            merged["section_padding_px"] = str(preset["section_padding_mm"])
        if "section_gap_px" not in preset and "section_gap_mm" in preset:
            merged["section_gap_px"] = str(preset["section_gap_mm"])
        merged["preset_name"] = str(name)
        normalized_presets[str(name)] = merged
    if active not in normalized_presets:
        active = default["preset_name"]
    return {
        "active": active,
        "presets": normalized_presets,
    }


def save_print_presets(active: str, presets: dict[str, dict[str, str]]) -> None:
    PRINT_PRESETS_FILE.write_text(
        json.dumps(
            {
                "active": active,
                "presets": presets,
            },
            indent=2,
        ),
        encoding="utf-8",
    )


def get_active_print_preset() -> dict[str, str]:
    state = load_print_presets()
    active = str(state["active"])
    presets = state["presets"]
    return dict(presets[active])


def build_print_preset_from_form(files: dict[str, tuple[str, bytes]]) -> dict[str, str]:
    base = default_print_preset()
    field_names = [
        "company_name",
        "department_label",
        "form_title",
        "date_label",
        "date_value",
        "section1_title",
        "section2_title",
        "section1_note",
        "section2_note",
        "header_bg",
        "status_header_bg",
        "branch_bg",
        "total_bg",
        "border_color",
        "accent_color",
        "page_padding_px",
        "print_page_margin_mm",
        "section_padding_px",
        "table_cell_padding_px",
        "section_gap_px",
        "signature_top_mm",
        "prepared_by_label",
        "prepared_by_name",
        "prepared_by_title",
        "verified_by_label",
        "verified_by_name",
        "verified_by_title",
    ]
    for field in field_names:
        value = files.get(field, ("", b""))[1].decode("utf-8", errors="ignore").strip()
        if value:
            base[field] = value
    preset_name = files.get("preset_name", ("", b""))[1].decode("utf-8", errors="ignore").strip() or base["preset_name"]
    base["preset_name"] = preset_name
    return base


def build_print_report_export_data() -> tuple[pd.DataFrame, pd.DataFrame]:
    queue = load_print_report_queue()
    status_breakdown = load_print_status_breakdown()

    updates_rows = []
    for item in queue:
        inactive = int(item.get("inactive_to_update", 0))
        active = int(item.get("active_to_update", 0))
        new_active = int(item.get("new_active_employees", 0))
        total = int(item.get("total", inactive + active + new_active))
        updates_rows.append(
            {
                "Branch": str(item.get("branch", "Not available")),
                "Inactive To Update": inactive,
                "Active To Update": active,
                "New Active Employees": new_active,
                "Grand Total": total,
            }
        )
    if updates_rows:
        updates_rows.append(
            {
                "Branch": "Grand Total",
                "Inactive To Update": sum(row["Inactive To Update"] for row in updates_rows),
                "Active To Update": sum(row["Active To Update"] for row in updates_rows),
                "New Active Employees": sum(row["New Active Employees"] for row in updates_rows),
                "Grand Total": sum(row["Grand Total"] for row in updates_rows),
            }
        )

    status_rows = []
    for item in status_breakdown:
        inactive = int(item.get("inactive", 0))
        active = int(item.get("active", 0))
        overall = int(item.get("overall", inactive + active))
        status_rows.append(
            {
                "Branch": str(item.get("branch", "Not available")),
                "Total Inactive Employees": inactive,
                "Total Active Employees": active,
                "Total Employees Overall": overall,
            }
        )
    if status_rows:
        status_rows.append(
            {
                "Branch": "Grand Total",
                "Total Inactive Employees": sum(row["Total Inactive Employees"] for row in status_rows),
                "Total Active Employees": sum(row["Total Active Employees"] for row in status_rows),
                "Total Employees Overall": sum(row["Total Employees Overall"] for row in status_rows),
            }
        )

    return pd.DataFrame(updates_rows), pd.DataFrame(status_rows)


def hex_to_argb(color: str) -> str:
    value = color.strip().lstrip("#")
    if len(value) == 3:
        value = "".join(char * 2 for char in value)
    if len(value) != 6:
        return "FF000000"
    return f"FF{value.upper()}"


def build_print_report_summary() -> dict[str, object]:
    queue = load_print_report_queue()
    status_breakdown = load_print_status_breakdown()

    update_rows: list[dict[str, int | str]] = []
    status_rows: list[dict[str, int | str]] = []
    grand_inactive = 0
    grand_active = 0
    grand_new = 0
    grand_new_in_system = 0
    grand_total = 0
    grand_total_inactive_employees = 0
    grand_total_active_employees = 0
    grand_total_employees_overall = 0

    for item in queue:
        inactive = int(item.get("inactive_to_update", 0))
        active = int(item.get("active_to_update", 0))
        new_active = int(item.get("new_active_employees", 0))
        new_in_system = int(item.get("new_in_system_since_last_month", 0))
        total = int(item.get("total", inactive + active + new_active + new_in_system))
        grand_inactive += inactive
        grand_active += active
        grand_new += new_active
        grand_new_in_system += new_in_system
        grand_total += total
        update_rows.append(
            {
                "Branch": str(item.get("branch", "Not available")),
                "Inactive To Update": inactive,
                "Active To Update": active,
                "New Active Employees": new_active,
                "New In System Since Last Snapshot": new_in_system,
                "Grand Total": total,
            }
        )

    for item in status_breakdown:
        inactive_total = int(item.get("inactive", 0))
        active_total = int(item.get("active", 0))
        overall_total = int(item.get("overall", inactive_total + active_total))
        grand_total_inactive_employees += inactive_total
        grand_total_active_employees += active_total
        grand_total_employees_overall += overall_total
        status_rows.append(
            {
                "Branch": str(item.get("branch", "Not available")),
                "Total Inactive Employees": inactive_total,
                "Total Active Employees": active_total,
                "Total Employees Overall": overall_total,
            }
        )

    return {
        "updates": update_rows,
        "status": status_rows,
        "updates_total": {
            "Branch": "Grand Total",
            "Inactive To Update": grand_inactive,
            "Active To Update": grand_active,
            "New Active Employees": grand_new,
            "New In System Since Last Snapshot": grand_new_in_system,
            "Grand Total": grand_total,
        },
        "status_total": {
            "Branch": "Grand Total",
            "Total Inactive Employees": grand_total_inactive_employees,
            "Total Active Employees": grand_total_active_employees,
            "Total Employees Overall": grand_total_employees_overall,
        },
    }


def export_print_report_workbook() -> bytes:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.page import PageMargins

    preset = get_active_print_preset()
    summary = build_print_report_summary()
    rendered_date_value = preset["date_value"] or format_local_timestamp(local_now())
    logo_path = get_print_logo_path()

    wb = Workbook()
    ws = wb.active
    ws.title = "Printable Form"

    line_color = preset["border_color"].replace("#", "").upper()
    line_side = Side(style="thin", color=line_color)
    medium_side = Side(style="medium", color=line_color)
    border = Border(left=line_side, right=line_side, top=line_side, bottom=line_side)
    outer_border = Border(left=medium_side, right=medium_side, top=medium_side, bottom=medium_side)
    header_fill = PatternFill("solid", fgColor=preset["header_bg"].replace("#", "").upper())
    status_fill = PatternFill("solid", fgColor=preset["status_header_bg"].replace("#", "").upper())
    branch_fill = PatternFill("solid", fgColor=preset["branch_bg"].replace("#", "").upper())
    total_fill = PatternFill("solid", fgColor=preset["total_bg"].replace("#", "").upper())
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def style_range(
        start_row: int,
        start_column: int,
        end_row: int,
        end_column: int,
        *,
        border_style: Border | None = None,
        fill_style: PatternFill | None = None,
        alignment_style: Alignment | None = None,
        font_style: Font | None = None,
    ) -> None:
        for row in ws.iter_rows(
            min_row=start_row,
            max_row=end_row,
            min_col=start_column,
            max_col=end_column,
        ):
            for cell in row:
                if border_style is not None:
                    cell.border = border_style
                if fill_style is not None:
                    cell.fill = fill_style
                if alignment_style is not None:
                    cell.alignment = alignment_style
                if font_style is not None:
                    cell.font = font_style

    def merge_and_style(
        start_row: int,
        start_column: int,
        end_row: int,
        end_column: int,
        value: object = "",
        *,
        border_style: Border | None = None,
        fill_style: PatternFill | None = None,
        alignment_style: Alignment | None = None,
        font_style: Font | None = None,
    ) -> None:
        ws.merge_cells(
            start_row=start_row,
            start_column=start_column,
            end_row=end_row,
            end_column=end_column,
        )
        anchor = ws.cell(start_row, start_column)
        anchor.value = value
        style_range(
            start_row,
            start_column,
            end_row,
            end_column,
            border_style=border_style,
            fill_style=fill_style,
            alignment_style=alignment_style,
            font_style=font_style,
        )

    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    margin_inch = float(preset["print_page_margin_mm"] or "10") / 25.4
    ws.page_margins = PageMargins(
        left=margin_inch,
        right=margin_inch,
        top=margin_inch,
        bottom=margin_inch,
        header=0.2,
        footer=0.2,
    )

    for col, width in {"A": 16, "B": 18, "C": 16, "D": 16, "E": 18, "F": 16}.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 23
    ws.row_dimensions[4].height = 17
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 30

    merge_and_style(1, 1, 5, 1, "", border_style=outer_border, alignment_style=center)
    if logo_path and logo_path.suffix.lower() in {".png", ".jpg", ".jpeg"}:
        try:
            logo_image = XLImage(str(logo_path))
            logo_image.width = 82
            logo_image.height = 82
            ws.add_image(logo_image, "A1")
        except Exception:
            ws["A1"] = preset["company_name"]
            ws["A1"].font = Font(size=14, bold=True)
    else:
        ws["A1"] = preset["company_name"]
        ws["A1"].font = Font(size=14, bold=True)

    header_rows = [
        ("B1:E1", preset["company_name"], 16, True),
        ("B2:E2", "Department:", 10, True),
        ("B3:E3", preset["department_label"], 13, True),
        ("B4:E4", "Form Title:", 10, True),
        ("B5:E5", preset["form_title"], 13, True),
    ]
    for cell_range, value, size, bold in header_rows:
        start_cell, end_cell = cell_range.split(":")
        merge_and_style(
            ws[start_cell].row,
            ws[start_cell].column,
            ws[end_cell].row,
            ws[end_cell].column,
            value,
            border_style=outer_border,
            alignment_style=center,
            font_style=Font(size=size, bold=bold),
        )

    merge_and_style(6, 1, 6, 2, preset["date_label"], border_style=outer_border, alignment_style=center, font_style=Font(size=11, bold=True))
    merge_and_style(6, 3, 6, 6, rendered_date_value, border_style=outer_border, alignment_style=center, font_style=Font(size=11, bold=True))

    start_row = 8
    ws.row_dimensions[start_row].height = 20
    merge_and_style(start_row, 1, start_row, 6, preset["section1_title"], border_style=border, alignment_style=left, font_style=Font(size=10, bold=True))
    note_row = start_row + 1
    ws.row_dimensions[note_row].height = 34
    merge_and_style(note_row, 1, note_row, 6, preset["section1_note"], border_style=border, alignment_style=left, font_style=Font(size=9))
    table_start = start_row + 2
    ws.row_dimensions[table_start].height = 24
    update_headers = ["Branch", "Inactive To Update", "Active To Update", "New Active Employees", "New In System", "Grand Total"]
    for index, header in enumerate(update_headers, start=1):
        cell = ws.cell(row=table_start, column=index, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = center
        cell.border = border

    current_row = table_start + 1
    update_rows = list(summary["updates"])
    if not update_rows:
        merge_and_style(current_row, 1, current_row, 6, "No branches added yet.", border_style=border, alignment_style=center)
        current_row += 1
    else:
        for row in update_rows:
            ws.row_dimensions[current_row].height = 22
            values = [
                row["Branch"],
                row["Inactive To Update"],
                row["Active To Update"],
                row["New Active Employees"],
                row["New In System Since Last Snapshot"],
                row["Grand Total"],
            ]
            for index, value in enumerate(values, start=1):
                cell = ws.cell(row=current_row, column=index, value=value)
                cell.border = border
                cell.alignment = center if index > 1 else left
                if index == 1:
                    cell.fill = branch_fill
                    cell.font = Font(bold=True)
            current_row += 1
    totals = summary["updates_total"]
    for index, value in enumerate(
        [
            totals["Branch"],
            totals["Inactive To Update"],
            totals["Active To Update"],
            totals["New Active Employees"],
            totals["New In System Since Last Snapshot"],
            totals["Grand Total"],
        ],
        start=1,
    ):
        cell = ws.cell(row=current_row, column=index, value=value)
        cell.fill = total_fill
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = center if index > 1 else left
    ws.row_dimensions[current_row].height = 22
    current_row += 2

    ws.row_dimensions[current_row].height = 20
    merge_and_style(current_row, 1, current_row, 6, preset["section2_title"], border_style=border, alignment_style=left, font_style=Font(size=10, bold=True))
    note_row = current_row + 1
    ws.row_dimensions[note_row].height = 34
    merge_and_style(note_row, 1, note_row, 6, preset["section2_note"], border_style=border, alignment_style=left, font_style=Font(size=9))
    current_row += 2

    status_headers = ["Branch", "Total Inactive Employees", "Total Active Employees", "Total Employees Overall"]
    ws.row_dimensions[current_row].height = 24
    merge_and_style(current_row, 1, current_row, 2, status_headers[0], border_style=border, fill_style=status_fill, alignment_style=center, font_style=Font(bold=True))
    for column, header in zip((3, 4, 5), status_headers[1:]):
        cell = ws.cell(row=current_row, column=column, value=header)
        cell.fill = status_fill
        cell.font = Font(bold=True)
        cell.alignment = center
        cell.border = border

    current_row += 1
    status_rows = list(summary["status"])
    if not status_rows:
        merge_and_style(current_row, 1, current_row, 6, "No updated system totals added yet.", border_style=border, alignment_style=center)
        current_row += 1
    else:
        for row in status_rows:
            ws.row_dimensions[current_row].height = 22
            merge_and_style(current_row, 1, current_row, 2, row["Branch"], border_style=border, fill_style=branch_fill, alignment_style=left, font_style=Font(bold=True))
            for column, value in zip((3, 4, 5), (
                row["Total Inactive Employees"],
                row["Total Active Employees"],
                row["Total Employees Overall"],
            )):
                cell = ws.cell(row=current_row, column=column, value=value)
                cell.border = border
                cell.alignment = center
            current_row += 1
    status_totals = summary["status_total"]
    ws.row_dimensions[current_row].height = 22
    merge_and_style(current_row, 1, current_row, 2, status_totals["Branch"], border_style=border, fill_style=total_fill, alignment_style=left, font_style=Font(bold=True))
    for column, value in zip((3, 4, 5), (
        status_totals["Total Inactive Employees"],
        status_totals["Total Active Employees"],
        status_totals["Total Employees Overall"],
    )):
        cell = ws.cell(row=current_row, column=column, value=value)
        cell.fill = total_fill
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = center

    current_row += 3
    prepared_row = current_row
    name_row = prepared_row + 2
    title_row = prepared_row + 3

    ws.merge_cells(start_row=prepared_row, start_column=1, end_row=prepared_row, end_column=2)
    ws.merge_cells(start_row=name_row, start_column=1, end_row=name_row, end_column=2)
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=2)
    ws[f"A{prepared_row}"] = preset["prepared_by_label"]
    ws[f"A{prepared_row}"].font = Font(bold=True)
    ws[f"A{name_row}"] = preset["prepared_by_name"]
    ws[f"A{name_row}"].font = Font(bold=True)
    ws[f"A{title_row}"] = preset["prepared_by_title"]
    ws[f"A{name_row}"].alignment = center
    ws[f"A{title_row}"].alignment = center

    ws.merge_cells(start_row=prepared_row, start_column=4, end_row=prepared_row, end_column=5)
    ws.merge_cells(start_row=name_row, start_column=4, end_row=name_row, end_column=5)
    ws.merge_cells(start_row=title_row, start_column=4, end_row=title_row, end_column=5)
    ws[f"D{prepared_row}"] = preset["verified_by_label"]
    ws[f"D{prepared_row}"].font = Font(bold=True)
    ws[f"D{name_row}"] = preset["verified_by_name"]
    ws[f"D{name_row}"].font = Font(bold=True)
    ws[f"D{title_row}"] = preset["verified_by_title"]
    ws[f"D{name_row}"].alignment = center
    ws[f"D{title_row}"].alignment = center
    ws.row_dimensions[prepared_row].height = 18
    ws.row_dimensions[name_row].height = 22
    ws.row_dimensions[title_row].height = 18

    ws.print_title_rows = "1:6"
    ws.print_area = "A1:F{}".format(title_row)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def export_print_report_pdf() -> bytes:
    chrome_path = find_chrome_executable()
    if chrome_path is None:
        raise RuntimeError(
            "Exact PDF export requires Google Chrome or Microsoft Edge installed on this computer."
        )

    html = render_print_report(embed_assets=True)
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        html_path = temp_path / "printable_form.html"
        pdf_path = temp_path / "printable_form_result.pdf"
        html_path.write_text(html, encoding="utf-8")

        chrome_command = [
            str(chrome_path),
            "--headless=new",
            "--disable-gpu",
            "--no-pdf-header-footer",
            f"--print-to-pdf={pdf_path}",
            html_path.as_uri(),
        ]
        result = subprocess.run(
            chrome_command,
            capture_output=True,
            text=True,
            check=False,
        )
        if result.returncode != 0 or not pdf_path.exists():
            raise RuntimeError(
                "Chrome could not generate the exact PDF export."
            )
        return pdf_path.read_bytes()


def find_chrome_executable() -> str | None:
    direct_candidates = [
        shutil.which("google-chrome"),
        shutil.which("google-chrome-stable"),
        shutil.which("chromium"),
        shutil.which("chromium-browser"),
        shutil.which("chrome.exe"),
        shutil.which("msedge.exe"),
    ]

    program_files = [
        os.environ.get("PROGRAMFILES", ""),
        os.environ.get("PROGRAMFILES(X86)", ""),
        "/mnt/c/Program Files",
        "/mnt/c/Program Files (x86)",
    ]
    path_candidates = []
    for base in program_files:
        if not base:
            continue
        path_candidates.extend(
            [
                Path(base) / "Google/Chrome/Application/chrome.exe",
                Path(base) / "Microsoft/Edge/Application/msedge.exe",
            ]
        )

    for candidate in direct_candidates:
        if candidate and Path(candidate).exists():
            return candidate
    for candidate in path_candidates:
        if candidate.exists():
            return str(candidate)
    return None


def print_queue_key(item: dict[str, object]) -> str:
    branch = str(item.get("branch", "")).strip().lower()
    if branch:
        return f"branch:{branch}"
    source_name = str(item.get("source_name", "")).strip().lower()
    if source_name:
        return f"source:{source_name}"
    return ""


def normalize_print_report_queue(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    seen: set[str] = set()
    normalized: list[dict[str, object]] = []
    for item in reversed(rows):
        key = print_queue_key(item)
        if key and key in seen:
            continue
        if key:
            seen.add(key)
        normalized.append(item)
    normalized.reverse()
    return normalized


def get_print_logo_path() -> Path | None:
    for suffix in PRINT_LOGO_EXTENSIONS:
        candidate = PRINT_LOGO_BASE.with_suffix(suffix)
        if candidate.exists():
            return candidate
    return None


def get_print_logo_src(embed_assets: bool = False) -> str:
    logo_path = get_print_logo_path()
    if not logo_path:
        return ""
    if not embed_assets:
        return f"/print-logo?name={quote(logo_path.name)}"

    mime_type = mimetypes.guess_type(logo_path.name)[0] or "application/octet-stream"
    encoded = base64.b64encode(logo_path.read_bytes()).decode("ascii")
    return f"data:{mime_type};base64,{encoded}"


def save_print_logo(file_info: tuple[str, bytes]) -> Path:
    filename, content = file_info
    suffix = Path(filename).suffix.lower()
    if suffix not in PRINT_LOGO_EXTENSIONS:
        raise ValueError("Logo must be a PNG, JPG, JPEG, SVG, or WEBP file.")

    for existing_suffix in PRINT_LOGO_EXTENSIONS:
        candidate = PRINT_LOGO_BASE.with_suffix(existing_suffix)
        if candidate.exists():
            candidate.unlink()

    target = PRINT_LOGO_BASE.with_suffix(suffix)
    target.write_bytes(content)
    return target


def add_current_report_to_print_queue(report_dir: Path) -> dict[str, object]:
    summary = summarize_reports(report_dir)
    queue = load_print_report_queue()
    entry = {
        "branch": summary["branch"] or "Not available",
        "inactive_to_update": int(summary["counts"].get("inactive_to_update.xlsx", 0)),
        "active_to_update": int(summary["counts"].get("active_to_update.xlsx", 0)),
        "new_active_employees": int(summary["counts"].get("new_active_employees.xlsx", 0)),
        "new_in_system_since_last_month": int(summary["counts"].get("new_in_system_since_last_month.xlsx", 0)),
        "total_inactive_employees": 0,
        "total_active_employees": 0,
        "total_employees_overall": 0,
        "total": int(summary["total"]),
        "generated_at": str(summary["generated_at"]),
        "source_name": str(summary["source_name"]),
        "system_file": str(summary["system_file"]),
        "hr_file": str(summary["hr_file"]),
    }
    entry_key = print_queue_key(entry)
    queue = [
        item for item in queue
        if print_queue_key(item) != entry_key or not entry_key
    ]
    queue.append(entry)
    save_print_report_queue(queue)
    return entry


def get_updated_system_path(report_dir: Path) -> Path | None:
    for filename in UPDATED_SYSTEM_FILENAMES:
        path = report_dir / filename
        if path.exists():
            return path
    return None


def get_updated_system_config_path(report_dir: Path) -> Path:
    return report_dir / UPDATED_SYSTEM_CONFIG_FILE


def load_updated_system_config(report_dir: Path) -> dict[str, str]:
    path = get_updated_system_config_path(report_dir)
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}
    if not isinstance(data, dict):
        return {}
    return {
        "status_column": str(data.get("status_column", "")).strip(),
        "branch_column": str(data.get("branch_column", "")).strip(),
    }


def save_updated_system_config(report_dir: Path, status_column: str, branch_column: str) -> None:
    path = get_updated_system_config_path(report_dir)
    path.write_text(
        json.dumps(
            {
                "status_column": status_column.strip(),
                "branch_column": branch_column.strip(),
            },
            indent=2,
        ),
        encoding="utf-8",
    )


def normalize_branch_name(value: object) -> str:
    return " ".join(str(value).strip().lower().split())


def compact_branch_name(value: object) -> str:
    return "".join(ch for ch in normalize_branch_name(value) if ch.isalnum())


def excel_column_letter(index: int) -> str:
    result = ""
    value = index + 1
    while value > 0:
        value, remainder = divmod(value - 1, 26)
        result = chr(65 + remainder) + result
    return result


def map_system_branch_label(value: object) -> str:
    normalized = normalize_branch_name(value)
    compact = compact_branch_name(value)
    if not compact:
        return UNMAPPED_BRANCH_LABEL

    for canonical, aliases in SYSTEM_BRANCH_GROUPS.items():
        for alias in aliases:
            normalized_alias = normalize_branch_name(alias)
            compact_alias = compact_branch_name(alias)
            if normalized_alias and normalized_alias in normalized:
                return canonical
            if compact_alias and compact_alias in compact:
                return canonical
            if compact and compact == compact_alias:
                return canonical
        
    return UNMAPPED_BRANCH_LABEL


def analyze_updated_system_file(
    path: Path,
    status_column_override: str = "",
    branch_column_override: str = "",
) -> dict[str, object]:
    # Headers are on row 1 and data starts on row 2.
    raw = read_excel_file(path, header=0)
    raw = normalize_columns(raw)
    raw = raw.rename(columns=DISPLAY_COLUMN_RENAMES)
    raw = raw.rename(columns={
        "employeeid": "id",
        "idno": "id",
        "status": "system_status",
        "systemstatus": "system_status",
    })

    if len(raw.columns) < 8:
        raise KeyError(
            f"Updated system file {path.name} must have at least 8 columns so column H can be used for branch mapping."
        )

    available_columns = list(raw.columns)
    column_lookup = {str(column).strip().lower(): column for column in available_columns}

    def resolve_column(
        preferred_names: set[str],
        override: str,
        fallback_index: int,
        *,
        allow_partial_match: bool = False,
    ) -> tuple[object, str]:
        override_key = override.strip().lower()
        if override_key and override_key in column_lookup:
            resolved = column_lookup[override_key]
            return resolved, "manual"
        for name in preferred_names:
            if name in column_lookup:
                return column_lookup[name], "auto"
        if allow_partial_match:
            for normalized_name, original_column in column_lookup.items():
                if any(name in normalized_name for name in preferred_names):
                    return original_column, "auto"
        return raw.columns[fallback_index], "fallback"

    status_column, status_source = resolve_column(
        UPDATED_SYSTEM_STATUS_COLUMN_ALIASES.union({"system_status"}),
        status_column_override,
        3,
        allow_partial_match=True,
    )
    branch_column, branch_source = resolve_column(
        UPDATED_SYSTEM_BRANCH_COLUMN_ALIASES.union({"branch_basis"}),
        branch_column_override,
        7,
        allow_partial_match=False,
    )

    system = raw.copy()
    if "id" not in system.columns:
        system["id"] = system.iloc[:, 0]
    system["system_status"] = system[status_column]
    system["branch_basis"] = system[branch_column]
    system["id"] = system["id"].astype(str).fillna("").str.strip()
    system["system_status"] = system["system_status"].astype(str).fillna("").str.strip().str.lower()
    system["branch_basis"] = system["branch_basis"].astype(str).fillna("").str.strip()
    system = system[system["id"] != ""].copy()
    return {
        "dataframe": system,
        "available_columns": available_columns,
        "status_column": str(status_column),
        "branch_column": str(branch_column),
        "status_column_source": status_source,
        "branch_column_source": branch_source,
        "status_column_letter": excel_column_letter(available_columns.index(status_column)),
        "branch_column_letter": excel_column_letter(available_columns.index(branch_column)),
    }


def summarize_updated_system(report_dir: Path) -> dict[str, object]:
    path = get_updated_system_path(report_dir)
    if path is None:
        return {
            "path": "",
            "active": 0,
            "inactive": 0,
            "overall": 0,
            "breakdown": [],
            "available_columns": [],
            "status_column": "",
            "branch_column": "",
            "status_column_source": "",
            "branch_column_source": "",
            "status_column_letter": "",
            "branch_column_letter": "",
        }

    config = load_updated_system_config(report_dir)
    analysis = analyze_updated_system_file(
        path,
        status_column_override=config.get("status_column", ""),
        branch_column_override=config.get("branch_column", ""),
    )
    system = analysis["dataframe"]
    statuses = system["system_status"].astype(str).str.strip().str.lower()
    active_count = int((statuses == "active").sum())
    overall = int(len(system.index))
    inactive_count = int(overall - active_count)
    breakdown_map: dict[str, dict[str, int | str]] = {}
    for _, row in system.iterrows():
        branch_label = map_system_branch_label(row.get("branch_basis", ""))
        entry = breakdown_map.setdefault(
            branch_label,
            {
                "branch": branch_label,
                "active": 0,
                "inactive": 0,
                "overall": 0,
            },
        )
        status = str(row.get("system_status", "")).strip().lower()
        entry["overall"] = int(entry["overall"]) + 1
        if status == "active":
            entry["active"] = int(entry["active"]) + 1
        else:
            entry["inactive"] = int(entry["inactive"]) + 1

    ordered_breakdown: list[dict[str, object]] = []
    for canonical in SYSTEM_BRANCH_GROUPS:
        ordered_breakdown.append(
            dict(
                breakdown_map.get(
                    canonical,
                    {
                        "branch": canonical,
                        "active": 0,
                        "inactive": 0,
                        "overall": 0,
                    },
                )
            )
        )
    if UNMAPPED_BRANCH_LABEL in breakdown_map and int(breakdown_map[UNMAPPED_BRANCH_LABEL]["overall"]) > 0:
        ordered_breakdown.append(dict(breakdown_map[UNMAPPED_BRANCH_LABEL]))

    return {
        "path": path.name,
        "active": active_count,
        "inactive": inactive_count,
        "overall": overall,
        "breakdown": ordered_breakdown,
        "available_columns": analysis["available_columns"],
        "status_column": analysis["status_column"],
        "branch_column": analysis["branch_column"],
        "status_column_source": analysis["status_column_source"],
        "branch_column_source": analysis["branch_column_source"],
        "status_column_letter": analysis["status_column_letter"],
        "branch_column_letter": analysis["branch_column_letter"],
    }


def upsert_status_summary_to_print_queue(report_dir: Path) -> dict[str, object]:
    status_summary = summarize_updated_system(report_dir)
    if not status_summary["path"]:
        raise ValueError("Upload an updated system_clean file first.")
    save_print_status_breakdown(status_summary["breakdown"])
    return status_summary


def upsert_filtered_report_to_print_queue(
    report_dir: Path,
    report_name: str,
    filtered_count: int,
) -> dict[str, object]:
    summary = summarize_reports(report_dir)
    queue = load_print_report_queue()
    branch = summary["branch"] or "Not available"
    source_name = str(summary["source_name"])
    entry_key = print_queue_key({"branch": branch, "source_name": source_name})
    filename_to_field = {
        "inactive_to_update.xlsx": "inactive_to_update",
        "active_to_update.xlsx": "active_to_update",
        "new_active_employees.xlsx": "new_active_employees",
        "new_in_system_since_last_month.xlsx": "new_in_system_since_last_month",
    }
    target_field = filename_to_field.get(report_name)
    if target_field is None:
        raise ValueError("Selected report cannot be added to the printable report.")

    existing_entry = None
    remaining_queue: list[dict[str, object]] = []
    for item in queue:
        if print_queue_key(item) == entry_key:
            existing_entry = item
            continue
        remaining_queue.append(item)

    entry = {
        "branch": branch,
        "inactive_to_update": 0,
        "active_to_update": 0,
        "new_active_employees": 0,
        "new_in_system_since_last_month": 0,
        "total": 0,
        "generated_at": str(summary["generated_at"]),
        "source_name": source_name,
        "system_file": str(summary["system_file"]),
        "hr_file": str(summary["hr_file"]),
    }
    if existing_entry:
        entry.update(existing_entry)

    entry[target_field] = int(filtered_count)
    entry["generated_at"] = str(summary["generated_at"])
    entry["source_name"] = source_name
    entry["system_file"] = str(summary["system_file"])
    entry["hr_file"] = str(summary["hr_file"])
    entry["total"] = (
        int(entry["inactive_to_update"])
        + int(entry["active_to_update"])
        + int(entry["new_active_employees"])
        + int(entry["new_in_system_since_last_month"])
    )

    remaining_queue.append(entry)
    save_print_report_queue(remaining_queue)
    return entry


def remove_print_report_entry(index: int) -> bool:
    queue = load_print_report_queue()
    if index < 0 or index >= len(queue):
        return False
    del queue[index]
    save_print_report_queue(queue)
    return True


def clear_print_report_queue() -> None:
    save_print_report_queue([])
    save_print_status_breakdown([])


def normalize_include_reports(include_reports: set[str] | None) -> set[str]:
    if include_reports is None:
        return set(REPORT_FILES)
    return {filename for filename in include_reports if filename in REPORT_FILES}


def include_query(include_reports: set[str] | None) -> str:
    normalized = normalize_include_reports(include_reports)
    if normalized == set(REPORT_FILES):
        return ""
    return "&include=" + quote(",".join(sorted(normalized)))


def latest_report_path(filename: str, report_dir: Path) -> Path | None:
    primary = report_dir / filename
    if primary.exists():
        return primary

    pattern = f"{primary.stem}_*{primary.suffix}"
    matches = sorted(
        report_dir.glob(pattern),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    return matches[0] if matches else None


def load_report(filename: str, report_dir: Path) -> pd.DataFrame:
    path = latest_report_path(filename, report_dir)
    if path is None:
        return pd.DataFrame()
    df = pd.read_excel(path, dtype=str).fillna("")
    if "id" in df.columns:
        df["id"] = df["id"].astype(str).str.strip()
    return df


def get_hr_file_path(report_dir: Path) -> Path | None:
    try:
        return find_input_file("hr_clean.xlsx", "hr_clean.xls", base_dir=report_dir)
    except FileNotFoundError:
        return None


def get_hr_branch_label(report_dir: Path) -> str:
    hr_path = get_hr_file_path(report_dir)
    if hr_path is None:
        return ""

    try:
        preview = read_excel_file(hr_path, header=None, nrows=1).fillna("")
    except Exception:
        return ""

    if preview.empty:
        return ""

    row_values = [str(value).strip() for value in preview.iloc[0].tolist()]
    non_empty = [value for value in row_values if value and value.lower() != "nan"]
    if not non_empty:
        return ""

    return " | ".join(non_empty)


def get_hr_row_label(report_dir: Path, row_number: int) -> str:
    hr_path = get_hr_file_path(report_dir)
    if hr_path is None:
        return ""

    try:
        preview = read_excel_file(hr_path, header=None, nrows=row_number + 1).fillna("")
    except Exception:
        return ""

    if preview.empty or len(preview.index) <= row_number:
        return ""

    row_values = [str(value).strip() for value in preview.iloc[row_number].tolist()]
    non_empty = [value for value in row_values if value and value.lower() != "nan"]
    return " | ".join(non_empty)


def summarize_reports(report_dir: Path, include_reports: set[str] | None = None) -> dict[str, object]:
    included = normalize_include_reports(include_reports)
    summary: dict[str, object] = {
        "branch": get_hr_branch_label(report_dir),
        "generated_at": format_local_timestamp(local_now()),
        "source_name": report_dir.name if report_dir != BASE_DIR else "project folder",
        "system_file": "",
        "hr_file": "",
        "counts": {},
        "total": 0,
        "included_reports": sorted(included),
    }

    system_path = latest_report_path("system_clean.xlsx", report_dir)
    if system_path is None:
        system_path = latest_report_path("system_clean.xls", report_dir)
    hr_path = get_hr_file_path(report_dir)

    summary["system_file"] = system_path.name if system_path else ""
    summary["hr_file"] = hr_path.name if hr_path else ""

    total = 0
    counts: dict[str, int] = {}
    for filename in REPORT_FILES:
        df = load_report(filename, report_dir)
        count = len(df.index) if filename in included else 0
        total += count
        counts[filename] = count

    summary["counts"] = counts
    summary["total"] = total
    return summary


def render_column_filter(df: pd.DataFrame, column_name: str, label: str) -> str:
    if column_name not in df.columns:
        return ""

    values = (
        df[column_name]
        .astype(str)
        .str.strip()
        .replace("nan", "")
    )
    options = sorted({value for value in values if value})
    option_tags = "".join(
        f'<option value="{escape(value)}"></option>'
        for value in options
    )

    return (
        '<div class="filter-box">'
        f'<label for="filter-{column_name}">{escape(label)}</label>'
        f'<div class="filter-control" data-filter-group="{escape(column_name)}">'
        f'<select id="filter-{column_name}-mode" data-column="{escape(column_name)}" data-role="mode">'
        '<option value="">All values</option>'
        '<option value="exact">Exact match</option>'
        '<option value="contains">Contains</option>'
        '<option value="not_contains">Does not contain</option>'
        '<option value="duplicates">Duplicates only</option>'
        "</select>"
        f'<div class="filter-values" data-column="{escape(column_name)}">'
        f'<input id="filter-{column_name}" data-column="{escape(column_name)}" data-role="filter-term" '
        f'type="text" list="filter-{column_name}-options" placeholder="Value 1">'
        "</div>"
        f'<button type="button" class="add-filter-value" data-column="{escape(column_name)}">Add Another Value</button>'
        f'<datalist id="filter-{column_name}-options">{option_tags}</datalist>'
        "</div>"
        "</div>"
    )


def render_table(df: pd.DataFrame) -> str:
    if df.empty:
        return "<p>No rows found.</p>"

    headers = '<th class="select-col">Mark</th>' + "".join(
        f"<th>{escape(str(col))}</th>" for col in df.columns
    )
    rows = []
    for index, row in df.iterrows():
        cells = "".join(f"<td>{escape(str(value))}</td>" for value in row.tolist())
        checkbox = (
            '<td class="select-col">'
            f'<input class="row-check" type="checkbox" aria-label="Highlight row {index + 1}">'
            "</td>"
        )
        rows.append(f"<tr>{checkbox}{cells}</tr>")
    body = "".join(rows)
    return f'<table id="report-table" class="report-table"><thead><tr>{headers}</tr></thead><tbody>{body}</tbody></table>'


def render_summary(df: pd.DataFrame) -> str:
    total_rows = len(df.index)
    active_count = 0
    inactive_count = 0

    if "hr_status" in df.columns:
        statuses = df["hr_status"].astype(str).str.strip().str.lower()
        active_count = int((statuses == "active").sum())
        inactive_count = int((statuses != "active").sum())

    cards = [
        ("Total Rows", str(total_rows)),
        ("HR Active", str(active_count)),
        ("HR Not Active", str(inactive_count)),
    ]

    return "".join(
        f'<div class="card"><div class="card-label">{escape(label)}</div>'
        f'<div class="card-value" data-summary-key="{escape(label.lower().replace(" ", "_"))}">{escape(value)}</div></div>'
        for label, value in cards
    )


def apply_single_filter(
    df: pd.DataFrame,
    column_name: str,
    mode: str,
    filter_values: list[str] | None = None,
) -> pd.DataFrame:
    if column_name not in df.columns or not mode:
        return df

    series = df[column_name].astype(str).fillna("").str.strip()
    normalized = series.str.lower()
    normalized_values = [
        item.strip().lower()
        for item in (filter_values or [])
        if item.strip()
    ]

    if mode == "duplicates":
        return df[normalized.ne("") & normalized.duplicated(keep=False)]
    if not normalized_values:
        return df
    if mode == "contains":
        mask = pd.Series(False, index=df.index)
        for filter_value in normalized_values:
            mask = mask | normalized.str.contains(filter_value, regex=False)
        return df[mask]
    if mode == "not_contains":
        mask = pd.Series(True, index=df.index)
        for filter_value in normalized_values:
            mask = mask & ~normalized.str.contains(filter_value, regex=False)
        return df[mask]
    mask = pd.Series(False, index=df.index)
    for filter_value in normalized_values:
        mask = mask | (normalized == filter_value)
    return df[mask]


def parse_filter_values(*raw_values: str) -> list[str]:
    values: list[str] = []
    for raw in raw_values:
        if not raw:
            continue
        for part in raw.replace("\r", "\n").split("\n"):
            cleaned = part.strip()
            if cleaned:
                values.append(cleaned)
    return values


def filter_report_dataframe(
    df: pd.DataFrame,
    search: str = "",
    department_mode: str = "",
    department_values: list[str] | None = None,
    position_mode: str = "",
    position_values: list[str] | None = None,
) -> pd.DataFrame:
    filtered = df.copy()
    search_value = search.strip().lower()

    if search_value:
        mask = filtered.astype(str).fillna("").apply(
            lambda row: search_value in " ".join(row).lower(),
            axis=1,
        )
        filtered = filtered[mask]

    filtered = apply_single_filter(
        filtered,
        "department",
        department_mode,
        department_values,
    )
    filtered = apply_single_filter(
        filtered,
        "position",
        position_mode,
        position_values,
    )
    return filtered


def export_report_dataframe(df: pd.DataFrame, filename_base: str, export_format: str) -> tuple[bytes, str, str]:
    safe_name = Path(filename_base).stem or "filtered_report"
    if export_format == "xlsx":
        output = BytesIO()
        df.to_excel(output, index=False)
        return (
            output.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f'{safe_name}_filtered.xlsx',
        )

    content = df.to_csv(index=False).encode("utf-8-sig")
    return (
        content,
        "text/csv; charset=utf-8",
        f"{safe_name}_filtered.csv",
    )


def parse_multipart(handler: BaseHTTPRequestHandler) -> dict[str, tuple[str, bytes]]:
    content_type = handler.headers.get("Content-Type", "")
    length = int(handler.headers.get("Content-Length", "0"))
    body = handler.rfile.read(length)

    raw_message = (
        f"Content-Type: {content_type}\r\n"
        f"MIME-Version: 1.0\r\n\r\n"
    ).encode("utf-8") + body
    message = BytesParser(policy=default).parsebytes(raw_message)

    files: dict[str, tuple[str, bytes]] = {}
    for part in message.iter_parts():
        name = part.get_param("name", header="content-disposition")
        filename = part.get_filename()
        if not name:
            continue
        files[name] = (
            Path(filename).name if filename else "",
            part.get_payload(decode=True) or b"",
        )
    return files


def save_uploaded_file(upload_name: str, file_info: tuple[str, bytes], run_dir: Path) -> Path:
    filename, content = file_info
    suffix = Path(filename).suffix.lower()
    target_name = f"{upload_name}{suffix}"
    target = run_dir / target_name
    target.write_bytes(content)
    return target


def render_app_tabs(active_tab: str, selected_report: str, include_reports: set[str] | None = None) -> str:
    include_part = include_query(include_reports)
    links = []
    for slug, label in APP_TABS.items():
        query = f"?tab={quote(slug)}&report={quote(selected_report)}{include_part}"
        current = " app-tab-active" if slug == active_tab else ""
        links.append(f'<a class="app-tab{current}" href="{query}">{escape(label)}</a>')
    return "".join(links)


def render_report_tabs(selected_report: str, include_reports: set[str] | None = None) -> str:
    include_part = include_query(include_reports)
    links = []
    for filename in REPORT_FILES:
        label = filename.replace(".xlsx", "").replace("_", " ").title()
        current = ' class="report-tab active"' if filename == selected_report else ' class="report-tab"'
        href = f"/?tab=compare-employees&report={quote(filename)}{include_part}"
        links.append(f'<a{current} href="{href}">{escape(label)}</a>')
    return "".join(links)


def render_select_options(options: list[str], selected_value: str = "", include_auto: bool = False) -> str:
    tags: list[str] = []
    if include_auto:
        selected = ' selected' if not selected_value else ""
        tags.append(f'<option value=""{selected}>Auto Detect</option>')
    for option in options:
        selected = ' selected' if option == selected_value else ""
        tags.append(f'<option value="{escape(option)}"{selected}>{escape(option)}</option>')
    return "".join(tags)


def render_compare_panel(
    selected_report: str,
    include_reports: set[str] | None = None,
    message: str = "",
    error: str = "",
) -> str:
    report_dir = get_active_report_dir()
    df = load_report(selected_report, report_dir)
    path = latest_report_path(selected_report, report_dir)
    report_name = path.name if path else selected_report
    data_source = report_dir.name if report_dir != BASE_DIR else "project folder"
    branch_label = get_hr_branch_label(report_dir)
    has_reports = any(latest_report_path(filename, report_dir) for filename in REPORT_FILES)
    queued_rows = load_print_report_queue()
    saved_runs = list_saved_report_runs()
    updated_system_error = ""
    try:
        updated_system_summary = summarize_updated_system(report_dir)
    except Exception as exc:
        updated_system_summary = {
            "path": "",
            "active": 0,
            "inactive": 0,
            "overall": 0,
            "breakdown": [],
        }
        updated_system_error = str(exc)

    download_links = []
    for filename in REPORT_FILES:
        report_path = latest_report_path(filename, report_dir)
        if report_path is None:
            continue
        label = filename.replace(".xlsx", "").replace("_", " ").title()
        href = f"/download?file={quote(report_path.name)}"
        download_links.append(
            f'<a class="download-link" href="{href}">Download {escape(label)}</a>'
        )

    add_to_print_html = ""
    if has_reports:
        add_to_print_html = f"""
          <form class="inline-form" method="post" enctype="multipart/form-data">
            <input type="hidden" name="action" value="add_print_report">
            <input type="hidden" name="report" value="{escape(selected_report)}">
            <button type="submit" class="download-link include-link">Add To Print Report</button>
          </form>
        """

    message_html = f'<div class="notice success">{escape(message)}</div>' if message else ""
    error_html = f'<div class="notice error">{escape(error)}</div>' if error else ""
    department_filter = render_column_filter(df, "department", "Department")
    position_filter = render_column_filter(df, "position", "Position")
    existing_run_options = ['<option value="">Select a previous comparison</option>']
    for run in saved_runs:
        selected = ' selected' if run["name"] == data_source else ""
        existing_run_options.append(
            f'<option value="{escape(run["name"])}"{selected}>'
            f'{escape(run["branch"])} | {escape(run["name"])} | {escape(run["updated_at"])}'
            "</option>"
        )
    available_updated_columns = [str(column_name) for column_name in updated_system_summary.get("available_columns", [])]
    status_column_label = ""
    branch_column_label = ""
    if updated_system_summary.get("status_column"):
        status_column_label = (
            f"{updated_system_summary['status_column']} "
            f"(Column {updated_system_summary['status_column_letter']}, {updated_system_summary['status_column_source']})"
        )
    if updated_system_summary.get("branch_column"):
        branch_column_label = (
            f"{updated_system_summary['branch_column']} "
            f"(Column {updated_system_summary['branch_column_letter']}, {updated_system_summary['branch_column_source']})"
        )
    workflow_guide = """
      <section class="panel workflow-panel">
        <div class="panel-title-row">
          <div>
            <div class="panel-kicker">Instructions</div>
            <h2>Monthly Workflow Guide</h2>
          </div>
        </div>
        <details class="workflow-details">
          <summary>Show step-by-step workflow</summary>
          <div class="workflow-content">
            <p><strong>Use this monthly order for each branch run.</strong></p>
            <ol>
              <li>Upload the branch <code>system_clean</code> and <code>hr_clean</code> in <strong>Run Comparison</strong>.</li>
              <li>Review the generated reports:
                <code>inactive_to_update</code>,
                <code>active_to_update</code>,
                <code>new_active_employees</code>,
                and <code>new_in_system_since_last_month</code>.
              </li>
              <li>Use <strong>Add To Print Report</strong> to send the branch counts into Section 1 of the printable form.</li>
              <li>Upload the branch <code>updated_system_clean</code> in <strong>Updated System Status</strong>.</li>
              <li>Review the active, inactive, and overall branch totals.</li>
              <li>Click <strong>Add Status Totals To Print Report</strong> to send those totals into Section 2 of the printable form.</li>
              <li>Repeat the same steps for the next branch.</li>
              <li>When all branches are complete, open the printable form, review it, then save/archive the final monthly report.</li>
              <li>After saving the final printable report, the live printable form can be reset for the next monthly cycle.</li>
            </ol>
            <p><strong>Important notes</strong></p>
            <ul>
              <li><code>new_in_system_since_last_month</code> follows the current run branch from the uploaded <code>hr_clean</code> context.</li>
              <li>That report uses <code>system_clean</code> only for the newly detected system employees.</li>
              <li>The first three reports use name details from <code>hr_clean</code>.</li>
              <li>The <strong>Updated System Status</strong> section is only for active, inactive, and overall totals. New employee ID detection stays in the main compare flow.</li>
              <li>If no previous snapshot exists for the branch yet, new-in-system counting is skipped for that run until the branch has a saved prior month baseline.</li>
            </ul>
          </div>
        </details>
      </section>
    """

    return f"""
      <section class="hero">
        <div class="eyebrow">Local Workflow</div>
        <h1>Compare Employees</h1>
        <p class="lede">Upload a fresh system file and HR file, rerun the comparison, review the results, and download the generated reports from one place.</p>
        <div class="hero-pills">
          <span>Browser Upload</span>
          <span>Excel Preview</span>
          <span>Download Ready</span>
        </div>
      </section>
      {workflow_guide}
      <section class="panel upload-panel">
        <div class="panel-title-row">
          <div>
            <div class="panel-kicker">Run Comparison</div>
            <h2>Upload Source Files</h2>
          </div>
          <div class="source-chip">Source: {escape(data_source)}</div>
        </div>
        <form method="post" enctype="multipart/form-data">
          <input type="hidden" name="tab" value="compare-employees">
          <input type="hidden" name="report" value="{escape(selected_report)}">
          <div class="upload-grid">
            <div>
              <label for="system_file">System file</label>
              <input id="system_file" name="system_file" type="file" accept=".xls,.xlsx" required>
            </div>
            <div>
              <label for="hr_file">HR file</label>
              <input id="hr_file" name="hr_file" type="file" accept=".xls,.xlsx" required>
            </div>
            <div class="button-cell">
              <button type="submit">Upload And Compare</button>
            </div>
          </div>
        </form>
        <form class="existing-run-form" method="post" enctype="multipart/form-data">
          <input type="hidden" name="action" value="load_existing_run">
          <input type="hidden" name="report" value="{escape(selected_report)}">
          <div class="existing-run-grid">
            <div>
              <label for="existing_run">Preview Existing Comparison</label>
              <select id="existing_run" name="existing_run">
                {''.join(existing_run_options)}
              </select>
            </div>
            <div class="button-cell">
              <button type="submit">Open Existing Preview</button>
            </div>
          </div>
        </form>
      </section>
      <section class="panel upload-panel">
        <div class="panel-title-row">
          <div>
            <div class="panel-kicker">Updated System Status</div>
            <h2>Upload Updated system_clean</h2>
          </div>
          <div class="source-chip">
            File: {escape(str(updated_system_summary["path"] or "Not uploaded"))}
          </div>
        </div>
        <p class="lede">This section calculates branch-based active, inactive, and overall employee totals from the updated system file using column H as the branch basis, then sends that branch breakdown to Section 2 of the printable report. Rows that do not match the defined branch aliases are counted under UNMAPPED / OTHER so the totals still balance.</p>
        {f'<div class="notice error">{escape(updated_system_error)}</div>' if updated_system_error else ''}
        <form method="post" enctype="multipart/form-data">
          <input type="hidden" name="action" value="upload_updated_system">
          <input type="hidden" name="report" value="{escape(selected_report)}">
          <div class="upload-grid">
            <div>
              <label for="updated_system_file">Updated system_clean file</label>
              <input id="updated_system_file" name="updated_system_file" type="file" accept=".xls,.xlsx" required>
            </div>
            <div>
              <label for="updated_status_column">Status Column</label>
              <select id="updated_status_column" name="updated_status_column">
                {render_select_options(available_updated_columns, str(updated_system_summary.get("status_column", "")), include_auto=True)}
              </select>
            </div>
            <div>
              <label for="updated_branch_column">Branch Column</label>
              <select id="updated_branch_column" name="updated_branch_column">
                {render_select_options(available_updated_columns, str(updated_system_summary.get("branch_column", "")), include_auto=True)}
              </select>
            </div>
            <div class="button-cell">
              <button type="submit">Upload Updated System</button>
            </div>
          </div>
        </form>
        {f'<div class="notice success">Detected status column: {escape(status_column_label)}. Detected branch column: {escape(branch_column_label)}.</div>' if status_column_label and branch_column_label and not updated_system_error else ''}
        <div class="summary">
          <div class="card"><div class="card-label">Total Inactive Employees</div><div class="card-value">{escape(str(updated_system_summary["inactive"]))}</div></div>
          <div class="card"><div class="card-label">Total Active Employees</div><div class="card-value">{escape(str(updated_system_summary["active"]))}</div></div>
          <div class="card"><div class="card-label">Total Employees Overall</div><div class="card-value">{escape(str(updated_system_summary["overall"]))}</div></div>
        </div>
        <div class="table-wrap">
          <table class="report-table">
            <thead>
              <tr>
                <th>Printable Branch</th>
                <th>Total Inactive Employees</th>
                <th>Total Active Employees</th>
                <th>Total Employees Overall</th>
              </tr>
            </thead>
            <tbody>
              {''.join(
                  f"<tr><td>{escape(str(item.get('branch', '')))}</td>"
                  f"<td>{escape(str(item.get('inactive', 0)))}</td>"
                  f"<td>{escape(str(item.get('active', 0)))}</td>"
                  f"<td>{escape(str(item.get('overall', 0)))}</td></tr>"
                  for item in updated_system_summary["breakdown"]
              ) or '<tr><td colspan="4" class="empty-state">No mapped branch totals found yet.</td></tr>'}
            </tbody>
          </table>
        </div>
        <form class="inline-form" method="post" enctype="multipart/form-data">
          <input type="hidden" name="action" value="add_status_summary_to_print_report">
          <input type="hidden" name="report" value="{escape(selected_report)}">
          <button type="submit" class="download-link include-link">Add Status Totals To Print Report</button>
        </form>
      </section>
      {message_html}
      {error_html}
      <section class="report-shell">
        <div class="panel-title-row">
          <div>
            <div class="panel-kicker">Results</div>
            {f'<div class="branch-chip">Branch: {escape(branch_label)}</div>' if branch_label else ''}
            <h2>{escape(report_name)}</h2>
          </div>
        </div>
        <div class="report-tabs">{render_report_tabs(selected_report)}</div>
        <div class="downloads">
          {''.join(download_links)}
          {add_to_print_html}
          <a class="download-link print-link" href="/print-report" target="_blank">View Printable Report ({len(queued_rows)})</a>
        </div>
        <div class="toolbar">
          <div class="search-box">
            <input id="report-search" type="search" placeholder="Search employee ID, name, address, or status">
          </div>
          {department_filter}
          {position_filter}
          <div class="filter-box export-box">
            <label for="export-format">Export Format</label>
            <div class="filter-control">
              <select id="export-format">
                <option value="csv">CSV</option>
                <option value="xlsx">XLSX</option>
              </select>
            </div>
          </div>
          <button id="export-filtered" type="button" class="download-link export-link" data-export-name="{escape(Path(report_name).stem)}" data-report-name="{escape(selected_report)}">Export Filtered Data</button>
          <button id="add-filtered-print" type="button" class="download-link export-link" data-report-name="{escape(selected_report)}">Add Filtered To Print Report</button>
          <div id="search-meta" class="search-meta">Showing all rows</div>
        </div>
        <form id="filtered-print-form" method="post" enctype="multipart/form-data" class="hidden-form">
          <input type="hidden" name="action" value="add_filtered_print_report">
          <input type="hidden" name="report" value="{escape(selected_report)}">
          <input type="hidden" name="search" value="">
          <input type="hidden" name="department_mode" value="">
          <input type="hidden" name="department_values" value="">
          <input type="hidden" name="position_mode" value="">
          <input type="hidden" name="position_values" value="">
        </form>
        <div class="summary">{render_summary(df)}</div>
        <div class="table-wrap">{render_table(df)}</div>
      </section>
    """


def render_stock_placeholder(selected_report: str) -> str:
    return f"""
      <section class="hero">
        <div class="eyebrow">Soon Development</div>
        <h1>Trace Duplicate Stock Items</h1>
        <p class="lede">This tab is reserved for the next module. The plan is to upload stock files, detect repeated item codes or descriptions, and show duplicate clusters with source references.</p>
        <div class="hero-pills">
          <span>Duplicate Detection</span>
          <span>Stock Audit</span>
          <span>Next Module</span>
        </div>
      </section>
      <section class="panel roadmap-grid">
        <div class="roadmap-card">
          <div class="panel-kicker">Planned Scope</div>
          <h2>What This Module Will Do</h2>
          <ul>
            <li>Upload stock inventory files from local folders.</li>
            <li>Detect duplicate item codes, names, and near-matches.</li>
            <li>Group duplicates into a reviewable dashboard.</li>
            <li>Export duplicate findings into Excel reports.</li>
          </ul>
        </div>
        <div class="roadmap-card accent-card">
          <div class="panel-kicker">Status</div>
          <h2>Not Yet Active</h2>
          <p>The dashboard tab is ready, but the duplicate stock comparison logic has not been built yet.</p>
          <a class="ghost-link" href="/?tab=compare-employees&report={quote(selected_report)}">Go Back To Compare Employees</a>
        </div>
      </section>
    """


def render_print_report(
    message: str = "",
    error: str = "",
    auto_print: bool = False,
    embed_assets: bool = False,
) -> str:
    queue = load_print_report_queue()
    status_breakdown = load_print_status_breakdown()
    report_dir = get_active_report_dir()
    preset_state = load_print_presets()
    active_preset_name = str(preset_state["active"])
    presets = preset_state["presets"]
    preset = dict(presets[active_preset_name])
    department_label = preset["department_label"]
    logo_url = get_print_logo_src(embed_assets=embed_assets)
    rows = []
    status_rows = []
    grand_inactive = 0
    grand_active = 0
    grand_new = 0
    grand_new_in_system = 0
    grand_total = 0
    grand_total_inactive_employees = 0
    grand_total_active_employees = 0
    grand_total_employees_overall = 0

    for item in queue:
        row_index = len(rows)
        inactive = int(item.get("inactive_to_update", 0))
        active = int(item.get("active_to_update", 0))
        new_active = int(item.get("new_active_employees", 0))
        new_in_system = int(item.get("new_in_system_since_last_month", 0))
        total = int(item.get("total", inactive + active + new_active + new_in_system))
        grand_inactive += inactive
        grand_active += active
        grand_new += new_active
        grand_new_in_system += new_in_system
        grand_total += total
        rows.append(
            "<tr>"
            f"<td class=\"branch-cell\">{escape(str(item.get('branch', 'Not available')))}</td>"
            f"<td class=\"count\">{inactive}</td>"
            f"<td class=\"count\">{active}</td>"
            f"<td class=\"count\">{new_active}</td>"
            f"<td class=\"count\">{new_in_system}</td>"
            f"<td class=\"count\">{total}</td>"
            "<td class=\"screen-only row-action-cell\">"
            "<form method=\"post\" enctype=\"multipart/form-data\">"
            "<input type=\"hidden\" name=\"action\" value=\"remove_print_row\">"
            f"<input type=\"hidden\" name=\"row_index\" value=\"{row_index}\">"
            "<button type=\"submit\" class=\"danger-button\">Remove</button>"
            "</form>"
            "</td>"
            "</tr>"
        )

    for item in status_breakdown:
        inactive_total = int(item.get("inactive", 0))
        active_total = int(item.get("active", 0))
        overall_total = int(item.get("overall", inactive_total + active_total))
        grand_total_inactive_employees += inactive_total
        grand_total_active_employees += active_total
        grand_total_employees_overall += overall_total
        status_rows.append(
            "<tr>"
            f"<td class=\"branch-cell\">{escape(str(item.get('branch', 'Not available')))}</td>"
            f"<td class=\"count\">{inactive_total}</td>"
            f"<td class=\"count\">{active_total}</td>"
            f"<td class=\"count\">{overall_total}</td>"
            "</tr>"
        )

    if not rows:
        rows.append(
            '<tr><td colspan="7" class="empty-state">No branches added yet.</td></tr>'
        )
    if not status_rows:
        status_rows.append(
            '<tr><td colspan="4" class="empty-state">No updated system totals added yet.</td></tr>'
        )

    rendered_date_value = preset["date_value"] or format_local_timestamp(local_now())

    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Employee Update Print Report</title>
  <style>
    :root {{
      --ink: #111827;
      --line: {escape(preset["border_color"])};
      --soft: #f3f4f6;
      --accent: {escape(preset["accent_color"])};
      --header-bg: {escape(preset["header_bg"])};
      --status-header-bg: {escape(preset["status_header_bg"])};
      --branch-bg: {escape(preset["branch_bg"])};
      --total-bg: {escape(preset["total_bg"])};
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: Arial, Helvetica, sans-serif;
      color: var(--ink);
      background: white;
    }}
    .page {{
      width: 210mm;
      max-width: 210mm;
      margin: 0 auto;
      padding: {escape(preset["page_padding_px"])}px;
    }}
    .actions {{
      display: flex;
      justify-content: flex-end;
      align-items: center;
      flex-wrap: wrap;
      gap: 10px;
      margin-bottom: 16px;
    }}
    .actions button, .actions a {{
      border: 1px solid var(--line);
      background: white;
      color: var(--ink);
      padding: 10px 14px;
      text-decoration: none;
      cursor: pointer;
      font-size: 14px;
    }}
    .logo-upload-form {{
      display: flex;
      align-items: center;
      gap: 8px;
      margin-right: auto;
    }}
    .logo-upload-form input[type="file"] {{
      max-width: 240px;
      font-size: 13px;
    }}
    .status-message {{
      margin-bottom: 16px;
      padding: 10px 12px;
      border: 1px solid #99f6e4;
      background: #ecfeff;
      color: #134e4a;
      font-size: 14px;
    }}
    .status-error {{
      border-color: #fecaca;
      background: #fef2f2;
      color: #991b1b;
    }}
    .sheet {{
      border: 2px solid var(--line);
      min-height: 277mm;
      display: flex;
      flex-direction: column;
      break-inside: avoid-page;
      page-break-inside: avoid;
    }}
    .sheet-header {{
      display: grid;
      grid-template-columns: 124px 1fr;
      border-bottom: 2px solid var(--line);
    }}
    .logo-box {{
      min-height: 138px;
      border-right: 2px solid var(--line);
      display: flex;
      align-items: center;
      justify-content: center;
      padding: 10px;
      background: white;
    }}
    .header-main {{
      display: grid;
      grid-template-rows: 38px 22px 31px 20px 27px;
    }}
    .header-cell {{
      border-bottom: 2px solid var(--line);
      padding: 2px 10px;
      display: flex;
      align-items: center;
    }}
    .header-cell:last-child {{
      border-bottom: 0;
    }}
    .company {{
      text-align: center;
      justify-content: center;
      font-size: 24px;
      font-weight: 700;
    }}
    .label {{
      font-weight: 700;
      margin-right: 6px;
    }}
    .department-value, .title {{
      text-align: center;
      justify-content: center;
      font-size: 18px;
      font-weight: 700;
    }}
    .meta-grid {{
      display: grid;
      grid-template-columns: 220px 1fr;
      border-bottom: 2px solid var(--line);
    }}
    .meta-cell {{
      padding: 12px;
      border-right: 2px solid var(--line);
      min-height: 58px;
      display: flex;
      align-items: center;
      font-weight: 700;
    }}
    .meta-cell:last-child {{
      border-right: 0;
      justify-content: center;
      text-align: center;
      font-weight: 600;
    }}
    .section {{
      padding: {escape(preset["section_padding_px"])}px;
      flex: 1;
      display: flex;
      flex-direction: column;
    }}
    .logo-svg {{
      width: 96px;
      height: 96px;
      display: block;
    }}
    .logo-image {{
      max-width: 96px;
      max-height: 96px;
      display: block;
      object-fit: contain;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      table-layout: fixed;
    }}
    th, td {{
      border: 1px solid var(--line);
      padding: {escape(preset["table_cell_padding_px"])}px 12px;
      text-align: left;
    }}
    th {{
      background: var(--header-bg);
    }}
    .branch-cell {{
      background: var(--branch-bg);
      font-weight: 700;
    }}
    .status-table th {{
      background: var(--status-header-bg);
    }}
    .count {{
      width: 14%;
      text-align: center;
      font-weight: 700;
    }}
    .total-row td {{
      font-weight: 700;
      background: var(--total-bg);
    }}
    .row-action-cell {{
      width: 120px;
      text-align: center;
    }}
    .row-action-cell form {{
      margin: 0;
    }}
    .danger-button {{
      border: 1px solid #b91c1c;
      background: #fff1f2;
      color: #b91c1c;
      padding: 8px 12px;
      cursor: pointer;
      font-size: 13px;
    }}
    .empty-state {{
      text-align: center;
      color: #6b7280;
      padding: 24px;
    }}
    .table-section-title {{
      margin: 0 0 10px;
      font-size: 14px;
      font-weight: 700;
      text-transform: uppercase;
      letter-spacing: 0.04em;
    }}
    .table-section-note {{
      margin: 0 0 12px;
      font-size: 12px;
      color: #4b5563;
      line-height: 1.5;
    }}
    .sub-table {{
      margin-top: {escape(preset["section_gap_px"])}px;
    }}
    .signature-row {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 48px;
      margin-top: auto;
      padding-top: {escape(preset["signature_top_mm"])}mm;
    }}
    .signature-block {{
      min-height: 96px;
    }}
    .signature-label {{
      font-weight: 700;
      margin-bottom: 28px;
    }}
    .signature-name {{
      font-weight: 700;
      text-transform: uppercase;
      margin-bottom: 4px;
    }}
    .signature-title {{
      text-transform: uppercase;
    }}
    @media print {{
      @page {{
        size: A4 portrait;
        margin: {escape(preset["print_page_margin_mm"])}mm;
      }}
      html {{
        -webkit-print-color-adjust: exact;
        print-color-adjust: exact;
      }}
      body, .sheet, .sheet-header, .logo-box, .header-cell, .meta-cell, table, thead, tbody, tr, th, td, .branch-cell, .total-row td {{
        -webkit-print-color-adjust: exact;
        print-color-adjust: exact;
      }}
      .actions {{ display: none; }}
      .screen-only {{ display: none; }}
      .page {{
        width: auto;
        max-width: none;
        padding: 0;
      }}
      body {{ background: white; }}
      .sheet {{
        min-height: calc(297mm - 20mm);
      }}
      .sheet-header {{
        grid-template-columns: 108px 1fr;
      }}
      .logo-box {{
        min-height: 118px;
        padding: 8px;
      }}
      .logo-svg, .logo-image {{
        max-width: 82px;
        max-height: 82px;
        width: 82px;
        height: 82px;
      }}
      .header-main {{
        grid-template-rows: 34px 20px 28px 18px 24px;
      }}
      .header-cell {{
        padding: 2px 8px;
      }}
      .company {{
        font-size: 20px;
      }}
      .department-value, .title {{
        font-size: 15px;
      }}
      .meta-cell {{
        min-height: 46px;
        padding: 8px 10px;
        font-size: 13px;
      }}
      .section {{
        padding: calc({escape(preset["section_padding_px"])}px * 0.7);
      }}
      .sub-table {{
        margin-top: calc({escape(preset["section_gap_px"])}px * 0.8);
      }}
      th, td {{
        padding: calc({escape(preset["table_cell_padding_px"])}px * 0.7) 8px;
        font-size: 12px;
      }}
      .signature-row {{
        gap: 24px;
        padding-top: calc({escape(preset["signature_top_mm"])}mm * 0.95);
      }}
      .signature-block {{
        min-height: 72px;
      }}
      .signature-label {{
        margin-bottom: 22px;
        font-size: 13px;
      }}
      .signature-name, .signature-title {{
        font-size: 12px;
      }}
      .table-section-note {{
        margin-bottom: 8px;
        font-size: 11px;
        line-height: 1.35;
      }}
    }}
  </style>
</head>
<body>
  <main class="page">
    <div class="actions">
      <form class="logo-upload-form" method="post" enctype="multipart/form-data">
        <input type="hidden" name="action" value="upload_logo">
        <input name="logo_file" type="file" accept=".png,.jpg,.jpeg,.svg,.webp" required>
        <button type="submit">Upload Logo</button>
      </form>
      <form method="post" enctype="multipart/form-data" class="screen-only">
        <input type="hidden" name="action" value="clear_print_queue">
        <button type="submit" class="danger-button">Clear Data</button>
      </form>
      <form method="post" enctype="multipart/form-data" class="screen-only">
        <input type="hidden" name="action" value="save_and_reset_print_queue">
        <button type="submit">Save Current Printable Form</button>
      </form>
      <a class="screen-only" href="/print-report-editor" target="_blank">Edit Printable Form</a>
      <a class="screen-only" href="/download-print-report?format=xlsx">Download Printable Excel</a>
      <a class="screen-only" href="/download-print-report?format=pdf">Download Printable PDF</a>
      <button type="button" onclick="window.print()">Print</button>
      <a href="/?tab=compare-employees&report=inactive_to_update.xlsx">Back To Dashboard</a>
    </div>
    {f'<div class="status-message">{escape(message)}</div>' if message else ''}
    {f'<div class="status-message status-error">{escape(error)}</div>' if error else ''}
    <section class="sheet">
      <div class="sheet-header">
        <div class="logo-box">
          {f'<img class="logo-image" src="{logo_url}" alt="QM Builders logo">' if logo_url else '''
          <svg class="logo-svg" viewBox="0 0 120 120" xmlns="http://www.w3.org/2000/svg" aria-label="QM Builders logo">
            <rect x="3" y="3" width="114" height="114" rx="14" fill="#ffffff" stroke="#e11d1d" stroke-width="5"/>
            <path d="M22 92 L44 20 H78 L96 92 Z" fill="#ef1c1c"/>
            <path d="M53 26 L69 26 L86 92 L70 92 Z" fill="#ffffff"/>
            <path d="M32 58 L46 40 L58 68 L42 68 Z" fill="#ffffff"/>
            <path d="M20 82 L24 82 L24 40 L20 40 Z" fill="#ef1c1c"/>
            <path d="M84 54 C84 44 92 38 100 38 C107 38 112 43 112 50 L112 91 L106 91 L106 52 C106 48 104 45 100 45 C95 45 91 49 91 55 L91 91 L84 91 Z" fill="#ef1c1c"/>
            <path d="M94 66 L108 66 L108 72 L94 72 Z" fill="#ffffff"/>
          </svg>
          '''}
        </div>
        <div class="header-main">
          <div class="header-cell company">{escape(preset["company_name"])}</div>
          <div class="header-cell"><span class="label">Department:</span></div>
          <div class="header-cell department-value">{escape(department_label)}</div>
          <div class="header-cell"><span class="label">Form Title:</span></div>
          <div class="header-cell title">{escape(preset["form_title"])}</div>
        </div>
      </div>
      <div class="meta-grid">
        <div class="meta-cell">{escape(preset["date_label"])}</div>
        <div class="meta-cell">{escape(rendered_date_value)}</div>
      </div>
      <div class="section">
        <div class="table-section-title">{escape(preset["section1_title"])}</div>
        <p class="table-section-note">{escape(preset["section1_note"])}</p>
        <table>
          <thead>
            <tr>
              <th>Branch</th>
              <th class="count">Inactive To Update</th>
              <th class="count">Active To Update</th>
              <th class="count">New Active Employees</th>
              <th class="count">New In System</th>
              <th class="count">Grand Total</th>
              <th class="screen-only">Action</th>
            </tr>
          </thead>
          <tbody>
            {''.join(rows)}
            <tr class="total-row">
              <td>Grand Total</td>
              <td class="count">{grand_inactive}</td>
              <td class="count">{grand_active}</td>
              <td class="count">{grand_new}</td>
              <td class="count">{grand_new_in_system}</td>
              <td class="count">{grand_total}</td>
            </tr>
          </tbody>
        </table>
        <div class="sub-table">
          <div class="table-section-title">{escape(preset["section2_title"])}</div>
          <p class="table-section-note">{escape(preset["section2_note"])}</p>
          <table class="status-table">
            <thead>
              <tr>
                <th>Branch</th>
                <th class="count">Total Inactive Employees</th>
                <th class="count">Total Active Employees</th>
                <th class="count">Total Employees Overall</th>
              </tr>
            </thead>
            <tbody>
              {''.join(status_rows)}
              <tr class="total-row">
                <td>Grand Total</td>
                <td class="count">{grand_total_inactive_employees}</td>
                <td class="count">{grand_total_active_employees}</td>
                <td class="count">{grand_total_employees_overall}</td>
              </tr>
            </tbody>
          </table>
        </div>
        <div class="signature-row">
          <div class="signature-block">
            <div class="signature-label">{escape(preset["prepared_by_label"])}</div>
            <div class="signature-name">{escape(preset["prepared_by_name"])}</div>
            <div class="signature-title">{escape(preset["prepared_by_title"])}</div>
          </div>
          <div class="signature-block">
            <div class="signature-label">{escape(preset["verified_by_label"])}</div>
            <div class="signature-name">{escape(preset["verified_by_name"])}</div>
            <div class="signature-title">{escape(preset["verified_by_title"])}</div>
          </div>
        </div>
      </div>
    </section>
  </main>
  {('<script>window.addEventListener("load", () => window.print());</script>' if auto_print else '')}
</body>
</html>"""


def render_print_report_editor(message: str = "", error: str = "") -> str:
    state = load_print_presets()
    active_preset_name = str(state["active"])
    presets = state["presets"]
    preset = dict(presets[active_preset_name])
    preset_options = "".join(
        f'<option value="{escape(name)}"{" selected" if name == active_preset_name else ""}>{escape(name)}</option>'
        for name in sorted(presets)
    )
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Printable Form Editor</title>
  <style>
    body {{
      margin: 0;
      font-family: Arial, Helvetica, sans-serif;
      background: #f8fafc;
      color: #111827;
    }}
    .page {{
      max-width: 1100px;
      margin: 0 auto;
      padding: 24px;
    }}
    .panel {{
      background: white;
      border: 1px solid #d1d5db;
      padding: 18px;
      margin-bottom: 16px;
    }}
    .grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
      gap: 12px;
      margin-bottom: 14px;
    }}
    label {{
      display: block;
      margin-bottom: 4px;
      font-size: 12px;
      font-weight: 700;
    }}
    input, select, button, a {{
      font: inherit;
    }}
    input, select {{
      width: 100%;
      border: 1px solid #cbd5e1;
      padding: 10px 12px;
      box-sizing: border-box;
    }}
    .actions {{
      display: flex;
      gap: 10px;
      flex-wrap: wrap;
      align-items: end;
      margin-bottom: 14px;
    }}
    .actions form {{
      margin: 0;
    }}
    button, .link-button {{
      border: 1px solid #1f2937;
      background: white;
      color: #111827;
      padding: 10px 14px;
      cursor: pointer;
      text-decoration: none;
    }}
    .message {{
      margin-bottom: 16px;
      padding: 10px 12px;
      border: 1px solid #99f6e4;
      background: #ecfeff;
      color: #134e4a;
    }}
    .error {{
      border-color: #fecaca;
      background: #fef2f2;
      color: #991b1b;
    }}
  </style>
</head>
<body>
  <main class="page">
    {f'<div class="message">{escape(message)}</div>' if message else ''}
    {f'<div class="message error">{escape(error)}</div>' if error else ''}
    <section class="panel">
      <div class="actions">
        <form method="post" enctype="multipart/form-data">
          <input type="hidden" name="action" value="apply_print_preset">
          <label for="apply_preset_name">Saved Preset</label>
          <select id="apply_preset_name" name="apply_preset_name">{preset_options}</select>
          <button type="submit">Apply Preset</button>
        </form>
        <a class="link-button" href="/print-report" target="_blank">Open Printable Report</a>
      </div>
      <form method="post" enctype="multipart/form-data">
        <input type="hidden" name="action" value="save_print_preset">
        <div class="grid">
          <div><label for="preset_name">Preset Name</label><input id="preset_name" name="preset_name" value="{escape(active_preset_name)}"></div>
          <div><label for="company_name">Company Name</label><input id="company_name" name="company_name" value="{escape(preset['company_name'])}"></div>
          <div><label for="department_label">Department</label><input id="department_label" name="department_label" value="{escape(preset['department_label'])}"></div>
          <div><label for="form_title">Form Title</label><input id="form_title" name="form_title" value="{escape(preset['form_title'])}"></div>
          <div><label for="date_label">Date Label</label><input id="date_label" name="date_label" value="{escape(preset['date_label'])}"></div>
          <div><label for="date_value">Date / Time Value</label><input id="date_value" name="date_value" value="{escape(preset['date_value'])}"></div>
          <div><label for="section1_title">Section 1 Title</label><input id="section1_title" name="section1_title" value="{escape(preset['section1_title'])}"></div>
          <div><label for="section2_title">Section 2 Title</label><input id="section2_title" name="section2_title" value="{escape(preset['section2_title'])}"></div>
          <div><label for="section1_note">Section 1 Note</label><input id="section1_note" name="section1_note" value="{escape(preset['section1_note'])}"></div>
          <div><label for="section2_note">Section 2 Note</label><input id="section2_note" name="section2_note" value="{escape(preset['section2_note'])}"></div>
          <div><label for="header_bg">Table Header Color</label><input id="header_bg" name="header_bg" type="color" value="{escape(preset['header_bg'])}"></div>
          <div><label for="status_header_bg">Status Header Color</label><input id="status_header_bg" name="status_header_bg" type="color" value="{escape(preset['status_header_bg'])}"></div>
          <div><label for="branch_bg">Branch Cell Color</label><input id="branch_bg" name="branch_bg" type="color" value="{escape(preset['branch_bg'])}"></div>
          <div><label for="total_bg">Total Row Color</label><input id="total_bg" name="total_bg" type="color" value="{escape(preset['total_bg'])}"></div>
          <div><label for="border_color">Border Color</label><input id="border_color" name="border_color" type="color" value="{escape(preset['border_color'])}"></div>
          <div><label for="accent_color">Accent Color</label><input id="accent_color" name="accent_color" type="color" value="{escape(preset['accent_color'])}"></div>
          <div><label for="page_padding_px">Screen Page Padding (px)</label><input id="page_padding_px" name="page_padding_px" type="number" step="0.1" value="{escape(preset['page_padding_px'])}"></div>
          <div><label for="print_page_margin_mm">Print Page Margin (mm)</label><input id="print_page_margin_mm" name="print_page_margin_mm" type="number" step="0.1" value="{escape(preset['print_page_margin_mm'])}"></div>
          <div><label for="section_padding_px">Section Padding (px)</label><input id="section_padding_px" name="section_padding_px" type="number" step="0.1" value="{escape(preset['section_padding_px'])}"></div>
          <div><label for="table_cell_padding_px">Table Cell Padding (px)</label><input id="table_cell_padding_px" name="table_cell_padding_px" type="number" step="0.1" value="{escape(preset['table_cell_padding_px'])}"></div>
          <div><label for="section_gap_px">Section Gap (px)</label><input id="section_gap_px" name="section_gap_px" type="number" step="0.1" value="{escape(preset['section_gap_px'])}"></div>
          <div><label for="signature_top_mm">Signature Top Spacing (mm)</label><input id="signature_top_mm" name="signature_top_mm" type="number" step="0.1" value="{escape(preset['signature_top_mm'])}"></div>
          <div><label for="prepared_by_label">Prepared Label</label><input id="prepared_by_label" name="prepared_by_label" value="{escape(preset['prepared_by_label'])}"></div>
          <div><label for="prepared_by_name">Prepared Name</label><input id="prepared_by_name" name="prepared_by_name" value="{escape(preset['prepared_by_name'])}"></div>
          <div><label for="prepared_by_title">Prepared Position</label><input id="prepared_by_title" name="prepared_by_title" value="{escape(preset['prepared_by_title'])}"></div>
          <div><label for="verified_by_label">Verified Label</label><input id="verified_by_label" name="verified_by_label" value="{escape(preset['verified_by_label'])}"></div>
          <div><label for="verified_by_name">Verified Name</label><input id="verified_by_name" name="verified_by_name" value="{escape(preset['verified_by_name'])}"></div>
          <div><label for="verified_by_title">Verified Position</label><input id="verified_by_title" name="verified_by_title" value="{escape(preset['verified_by_title'])}"></div>
        </div>
        <button type="submit">Save And Apply Preset</button>
      </form>
    </section>
  </main>
</body>
</html>"""


def render_page(
    active_tab: str,
    selected_report: str,
    include_reports: set[str] | None = None,
    message: str = "",
    error: str = "",
) -> str:
    if active_tab not in APP_TABS:
        active_tab = "compare-employees"
    if selected_report not in REPORT_FILES:
        selected_report = REPORT_FILES[0]

    if active_tab == "trace-duplicate-stock-items":
        content = render_stock_placeholder(selected_report)
    else:
        content = render_compare_panel(
            selected_report,
            include_reports=include_reports,
            message=message,
            error=error,
        )

    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Operations Dashboard</title>
  <style>
    :root {{
      --bg: #f6efe4;
      --paper: #fffaf2;
      --ink: #16202a;
      --muted: #6c7380;
      --line: #d8cab9;
      --accent: #0f766e;
      --accent-2: #c96f2d;
      --accent-soft: #ddf3ef;
      --sand: #f0e4d2;
      --good: #eaf8ef;
      --good-ink: #166534;
      --bad: #feeeee;
      --bad-ink: #9f1239;
      --shadow: 0 24px 70px rgba(73, 48, 18, 0.12);
    }}
    * {{
      box-sizing: border-box;
    }}
    body {{
      margin: 0;
      font-family: "Segoe UI", Tahoma, Geneva, Verdana, sans-serif;
      color: var(--ink);
      background:
        radial-gradient(circle at top left, rgba(201, 111, 45, 0.18), transparent 32%),
        radial-gradient(circle at top right, rgba(15, 118, 110, 0.18), transparent 28%),
        linear-gradient(180deg, #efe3d0 0%, var(--bg) 55%, #f8f3eb 100%);
      min-height: 100vh;
    }}
    .page-shell {{
      max-width: 1240px;
      margin: 0 auto;
      padding: 28px 20px 48px;
      animation: fade-up 500ms ease;
    }}
    .masthead {{
      display: grid;
      grid-template-columns: 1.2fr 0.8fr;
      gap: 18px;
      margin-bottom: 20px;
    }}
    .mast-card {{
      background: rgba(255, 250, 242, 0.8);
      backdrop-filter: blur(6px);
      border: 1px solid rgba(216, 202, 185, 0.9);
      border-radius: 22px;
      padding: 22px;
      box-shadow: var(--shadow);
    }}
    .brand-line {{
      display: flex;
      justify-content: space-between;
      gap: 12px;
      align-items: center;
      margin-bottom: 18px;
    }}
    .brand {{
      letter-spacing: 0.12em;
      text-transform: uppercase;
      font-size: 12px;
      color: var(--muted);
    }}
    .pulse {{
      width: 12px;
      height: 12px;
      border-radius: 999px;
      background: var(--accent);
      box-shadow: 0 0 0 rgba(15, 118, 110, 0.4);
      animation: pulse 1.8s infinite;
    }}
    .mast-card h1 {{
      margin: 0 0 10px;
      font-family: Georgia, "Times New Roman", serif;
      font-size: clamp(32px, 4vw, 54px);
      line-height: 0.95;
    }}
    .mast-card p {{
      margin: 0;
      color: var(--muted);
      font-size: 16px;
      line-height: 1.5;
    }}
    .stat-grid {{
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 12px;
      height: 100%;
    }}
    .stat-chip {{
      background: linear-gradient(180deg, #fff, #f8f0e4);
      border: 1px solid var(--line);
      border-radius: 18px;
      padding: 16px;
      display: flex;
      flex-direction: column;
      justify-content: end;
      animation: float-in 700ms ease both;
    }}
    .stat-chip:nth-child(2) {{ animation-delay: 90ms; }}
    .stat-chip:nth-child(3) {{ animation-delay: 160ms; }}
    .stat-chip:nth-child(4) {{ animation-delay: 230ms; }}
    .stat-label {{
      color: var(--muted);
      font-size: 12px;
      text-transform: uppercase;
      letter-spacing: 0.08em;
    }}
    .stat-value {{
      font-size: 22px;
      margin-top: 10px;
    }}
    .app-tabs {{
      display: flex;
      gap: 12px;
      flex-wrap: wrap;
      margin-bottom: 18px;
    }}
    .app-tab {{
      text-decoration: none;
      color: var(--ink);
      background: rgba(255, 250, 242, 0.84);
      border: 1px solid var(--line);
      border-radius: 999px;
      padding: 12px 18px;
      font-size: 15px;
      transition: transform 180ms ease, background 180ms ease;
    }}
    .app-tab:hover {{
      transform: translateY(-1px);
    }}
    .app-tab-active {{
      background: var(--accent);
      color: white;
      border-color: var(--accent);
    }}
    .hero, .panel, .report-shell {{
      background: rgba(255, 250, 242, 0.82);
      border: 1px solid rgba(216, 202, 185, 0.95);
      border-radius: 22px;
      padding: 22px;
      box-shadow: var(--shadow);
      margin-bottom: 18px;
      animation: fade-up 600ms ease both;
    }}
    .hero {{
      position: relative;
      overflow: hidden;
    }}
    .hero::after {{
      content: "";
      position: absolute;
      inset: auto -60px -60px auto;
      width: 220px;
      height: 220px;
      border-radius: 999px;
      background: radial-gradient(circle, rgba(15, 118, 110, 0.16), transparent 65%);
      pointer-events: none;
    }}
    .eyebrow, .panel-kicker {{
      color: var(--accent-2);
      text-transform: uppercase;
      letter-spacing: 0.12em;
      font-size: 12px;
      margin-bottom: 10px;
    }}
    .hero h1, .panel h2, .report-shell h2 {{
      margin: 0 0 10px;
      font-family: Georgia, "Times New Roman", serif;
    }}
    .lede {{
      margin: 0;
      max-width: 760px;
      color: var(--muted);
      line-height: 1.6;
      font-size: 16px;
    }}
    .hero-pills {{
      display: flex;
      gap: 10px;
      flex-wrap: wrap;
      margin-top: 18px;
    }}
    .hero-pills span {{
      border: 1px solid var(--line);
      background: white;
      border-radius: 999px;
      padding: 8px 12px;
      font-size: 13px;
    }}
    .workflow-details {{
      border: 1px solid rgba(216, 202, 185, 0.95);
      border-radius: 16px;
      background: rgba(255, 255, 255, 0.7);
      overflow: hidden;
    }}
    .workflow-details summary {{
      cursor: pointer;
      padding: 14px 16px;
      font-weight: 700;
      list-style: none;
    }}
    .workflow-details summary::-webkit-details-marker {{
      display: none;
    }}
    .workflow-content {{
      padding: 0 16px 16px;
      color: var(--ink);
      line-height: 1.6;
    }}
    .workflow-content p {{
      margin: 0 0 10px;
    }}
    .workflow-content ol,
    .workflow-content ul {{
      margin: 0 0 12px 20px;
      padding: 0;
    }}
    .workflow-content li {{
      margin-bottom: 8px;
    }}
    .panel-title-row {{
      display: flex;
      justify-content: space-between;
      align-items: start;
      gap: 14px;
      margin-bottom: 18px;
    }}
    .source-chip {{
      background: var(--sand);
      border: 1px solid var(--line);
      border-radius: 999px;
      padding: 10px 12px;
      font-size: 13px;
      color: var(--muted);
    }}
    .branch-chip {{
      display: inline-flex;
      align-items: center;
      margin-bottom: 10px;
      background: rgba(15, 118, 110, 0.08);
      border: 1px solid rgba(15, 118, 110, 0.18);
      border-radius: 999px;
      padding: 8px 12px;
      font-size: 13px;
      color: var(--accent);
    }}
    .upload-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
      gap: 14px;
      align-items: end;
    }}
    label {{
      display: block;
      margin-bottom: 8px;
      color: var(--muted);
      font-size: 14px;
    }}
    input[type="file"] {{
      width: 100%;
      border: 1px solid var(--line);
      background: #fffdf8;
      border-radius: 12px;
      padding: 10px;
      font-family: inherit;
    }}
    button {{
      width: 100%;
      border: 0;
      border-radius: 14px;
      background: linear-gradient(135deg, var(--accent), #0b5c56);
      color: white;
      padding: 14px 16px;
      font-size: 15px;
      cursor: pointer;
      box-shadow: 0 14px 30px rgba(15, 118, 110, 0.18);
    }}
    .button-cell {{
      align-self: stretch;
    }}
    .existing-run-form {{
      margin-top: 16px;
      padding-top: 16px;
      border-top: 1px solid var(--line);
    }}
    .existing-run-grid {{
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 14px;
      align-items: end;
    }}
    .report-tabs, .downloads {{
      display: flex;
      gap: 10px;
      flex-wrap: wrap;
      margin-bottom: 16px;
    }}
    .inline-form {{
      margin: 0;
    }}
    .hidden-form {{
      display: none;
    }}
    .toolbar {{
      display: flex;
      gap: 12px;
      flex-wrap: wrap;
      align-items: center;
      margin-bottom: 16px;
    }}
    .search-box {{
      min-width: 260px;
      flex: 1 1 320px;
    }}
    .search-box input, .filter-box select, .filter-box input, .existing-run-grid select {{
      width: 100%;
      border: 1px solid var(--line);
      background: #fffdf8;
      border-radius: 12px;
      padding: 11px 14px;
      font-family: "Segoe UI", Tahoma, Geneva, Verdana, sans-serif;
      font-size: 14px;
    }}
    .filter-box {{
      min-width: 180px;
      flex: 0 1 220px;
    }}
    .export-box {{
      flex-basis: 160px;
    }}
    .filter-control {{
      display: grid;
      gap: 8px;
    }}
    .filter-values {{
      display: grid;
      gap: 8px;
    }}
    .add-filter-value {{
      width: auto;
      border: 1px solid var(--line);
      border-radius: 10px;
      background: #fffdf8;
      color: var(--accent);
      padding: 10px 12px;
      font-size: 13px;
      box-shadow: none;
      justify-self: start;
    }}
    .search-meta {{
      color: var(--muted);
      font-size: 13px;
    }}
    .report-tab, .download-link, .ghost-link {{
      text-decoration: none;
      border: 1px solid var(--line);
      border-radius: 999px;
      padding: 10px 14px;
      font-size: 14px;
      background: white;
      color: var(--ink);
    }}
    button.download-link {{
      font: inherit;
      cursor: pointer;
    }}
    .export-link {{
      width: auto;
      box-shadow: none;
    }}
    .report-tab.active {{
      background: var(--accent-soft);
      border-color: rgba(15, 118, 110, 0.25);
    }}
    .download-link {{
      color: var(--accent);
    }}
    .ghost-link {{
      display: inline-block;
      margin-top: 14px;
    }}
    .notice {{
      border-radius: 16px;
      padding: 14px 16px;
      margin-bottom: 14px;
      font-size: 14px;
      box-shadow: var(--shadow);
    }}
    .notice.success {{
      background: var(--good);
      color: var(--good-ink);
    }}
    .notice.error {{
      background: var(--bad);
      color: var(--bad-ink);
    }}
    .summary {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
      gap: 12px;
      margin-bottom: 16px;
    }}
    .card {{
      background: linear-gradient(180deg, #fff, #faf3e9);
      border: 1px solid var(--line);
      border-radius: 18px;
      padding: 16px;
      animation: float-in 700ms ease both;
    }}
    .card-label {{
      color: var(--muted);
      font-size: 13px;
      margin-bottom: 8px;
    }}
    .card-value {{
      font-size: 30px;
      font-weight: 700;
      line-height: 1;
    }}
    .table-wrap {{
      overflow: auto;
      background: white;
      border: 1px solid var(--line);
      border-radius: 18px;
      max-height: 520px;
    }}
    .report-table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 14px;
      font-family: "Segoe UI", Tahoma, Geneva, Verdana, sans-serif;
    }}
    .report-table th, .report-table td {{
      border-bottom: 1px solid #eee2cf;
      padding: 10px 12px;
      text-align: left;
      vertical-align: top;
      white-space: nowrap;
    }}
    .report-table tbody tr.row-selected {{
      background: #fff4cf;
    }}
    .report-table th {{
      position: sticky;
      top: 0;
      background: var(--accent-soft);
      font-weight: 600;
      letter-spacing: 0.01em;
    }}
    .select-col {{
      width: 72px;
      min-width: 72px;
      text-align: center;
    }}
    .row-check {{
      width: 18px;
      height: 18px;
      accent-color: var(--accent);
      cursor: pointer;
    }}
    .report-table td:first-child, .report-table th:first-child {{
      font-family: inherit;
    }}
    .report-table td:nth-child(2), .report-table th:nth-child(2) {{
      font-family: "Consolas", "Courier New", monospace;
    }}
    p {{
      margin: 0;
    }}
    .roadmap-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
      gap: 18px;
    }}
    .roadmap-card {{
      background: white;
      border: 1px solid var(--line);
      border-radius: 18px;
      padding: 20px;
    }}
    .accent-card {{
      background: linear-gradient(180deg, #fff5ea, #fff);
    }}
    .roadmap-card ul {{
      margin: 0;
      padding-left: 18px;
      color: var(--muted);
      line-height: 1.7;
    }}
    @keyframes fade-up {{
      from {{ opacity: 0; transform: translateY(18px); }}
      to {{ opacity: 1; transform: translateY(0); }}
    }}
    @keyframes float-in {{
      from {{ opacity: 0; transform: translateY(14px) scale(0.98); }}
      to {{ opacity: 1; transform: translateY(0) scale(1); }}
    }}
    @keyframes pulse {{
      0% {{ box-shadow: 0 0 0 0 rgba(15, 118, 110, 0.4); }}
      70% {{ box-shadow: 0 0 0 14px rgba(15, 118, 110, 0); }}
      100% {{ box-shadow: 0 0 0 0 rgba(15, 118, 110, 0); }}
    }}
    @media (max-width: 900px) {{
      .masthead {{
        grid-template-columns: 1fr;
      }}
      .panel-title-row {{
        flex-direction: column;
      }}
      .existing-run-grid {{
        grid-template-columns: 1fr;
      }}
    }}
  </style>
</head>
<body>
  <main class="page-shell">
    <section class="masthead">
      <div class="mast-card">
        <div class="brand-line">
          <div class="brand">Operations Dashboard</div>
          <div class="pulse"></div>
        </div>
        <h1>Flexible local tools for operations review.</h1>
        <p>Start with employee comparison today, then extend the same dashboard into stock audit and duplicate tracing next.</p>
      </div>
      <div class="stat-grid">
        <div class="stat-chip">
          <div class="stat-label">Current Module</div>
          <div class="stat-value">Employee Comparison</div>
        </div>
        <div class="stat-chip">
          <div class="stat-label">Preview Mode</div>
          <div class="stat-value">Local Browser App</div>
        </div>
        <div class="stat-chip">
          <div class="stat-label">Next Module</div>
          <div class="stat-value">Duplicate Stock Trace</div>
        </div>
        <div class="stat-chip">
          <div class="stat-label">File Support</div>
          <div class="stat-value">XLS And XLSX</div>
        </div>
      </div>
    </section>
    <nav class="app-tabs">{render_app_tabs(active_tab, selected_report, include_reports)}</nav>
    {content}
  </main>
<script>
  (function () {{
    const input = document.getElementById("report-search");
    const table = document.getElementById("report-table");
    const meta = document.getElementById("search-meta");
    const exportButton = document.getElementById("export-filtered");
    const addFilteredPrintButton = document.getElementById("add-filtered-print");
    const filteredPrintForm = document.getElementById("filtered-print-form");
    const exportFormat = document.getElementById("export-format");
    if (!input || !table || !meta) {{
      return;
    }}

    const filterModes = Array.from(document.querySelectorAll('[data-role="mode"]'));
    const addFilterButtons = Array.from(document.querySelectorAll(".add-filter-value"));
    const filterPairs = filterModes.map((modeInput) => {{
      const columnName = (modeInput.dataset.column || "").trim().toLowerCase();
      const group = document.querySelector('[data-filter-group="' + columnName + '"]');
      const valuesContainer = group?.querySelector(".filter-values") || null;
      return {{
        columnName,
        modeInput,
        group,
        valuesContainer,
      }};
    }});
    const headers = Array.from(table.querySelectorAll("thead th")).map((cell) =>
      cell.textContent.trim().toLowerCase()
    );
    const rows = Array.from(table.querySelectorAll("tbody tr"));
    const duplicateMaps = new Map();
    const summaryNodes = {{
      totalRows: document.querySelector('[data-summary-key="total_rows"]'),
      hrActive: document.querySelector('[data-summary-key="hr_active"]'),
      hrNotActive: document.querySelector('[data-summary-key="hr_not_active"]'),
    }};
    const hrStatusIndex = headers.indexOf("hr_status");
    const dataHeaders = headers.slice(1);

    filterPairs.forEach((filter) => {{
      const columnName = filter.columnName;
      const columnIndex = headers.indexOf(columnName);
      if (columnIndex === -1) {{
        return;
      }}

      const counts = new Map();
      rows.forEach((row) => {{
        const cells = row.querySelectorAll("td");
        const value = (cells[columnIndex]?.textContent || "").trim();
        if (!value) {{
          return;
        }}
        counts.set(value, (counts.get(value) || 0) + 1);
      }});
      duplicateMaps.set(columnName, counts);
    }});

    const getVisibleRows = () => rows.filter((row) => row.style.display !== "none");
    const getFilterValues = (filter) => Array.from(
      filter.valuesContainer?.querySelectorAll('[data-role="filter-term"]') || []
    )
      .map((inputNode) => inputNode.value.trim())
      .filter(Boolean);

    const writeFilterStateToForm = () => {{
      if (!filteredPrintForm) {{
        return;
      }}
      const setValue = (name, value) => {{
        const field = filteredPrintForm.querySelector('[name="' + name + '"]');
        if (field) {{
          field.value = value;
        }}
      }};
      setValue("search", input.value.trim());
      filterPairs.forEach((filter) => {{
        setValue(filter.columnName + "_mode", (filter.modeInput?.value || "").trim());
        setValue(filter.columnName + "_values", getFilterValues(filter).join("\\n"));
      }});
    }};

    const updateSummary = (visibleRows) => {{
      if (summaryNodes.totalRows) {{
        summaryNodes.totalRows.textContent = String(visibleRows.length);
      }}

      if (hrStatusIndex === -1) {{
        return;
      }}

      let activeCount = 0;
      let inactiveCount = 0;
      visibleRows.forEach((row) => {{
        const cells = row.querySelectorAll("td");
        const status = (cells[hrStatusIndex]?.textContent || "").trim().toLowerCase();
        if (status === "active") {{
          activeCount += 1;
        }} else {{
          inactiveCount += 1;
        }}
      }});

      if (summaryNodes.hrActive) {{
        summaryNodes.hrActive.textContent = String(activeCount);
      }}
      if (summaryNodes.hrNotActive) {{
        summaryNodes.hrNotActive.textContent = String(inactiveCount);
      }}
    }};

    const exportFilteredRows = () => {{
      const params = new URLSearchParams();
      params.set("report", (exportButton?.dataset.reportName || "").trim());
      params.set("format", (exportFormat?.value || "csv").trim());
      params.set("search", input.value.trim());

      filterPairs.forEach((filter) => {{
        const mode = (filter.modeInput?.value || "").trim();
        if (!filter.columnName) {{
          return;
        }}
        params.set(filter.columnName + "_mode", mode);
        params.set(filter.columnName + "_values", getFilterValues(filter).join("\\n"));
      }});

      window.location.href = "/export-filtered?" + params.toString();
    }};

    const update = () => {{
      const query = input.value.trim().toLowerCase();

      rows.forEach((row) => {{
        const text = row.textContent.toLowerCase();
        const cells = Array.from(row.querySelectorAll("td"));
        const queryMatch = !query || text.includes(query);

        const filterMatch = filterPairs.every((filter) => {{
          const columnName = filter.columnName;
          const columnIndex = headers.indexOf(columnName);
          if (columnIndex === -1) {{
            return true;
          }}

          const mode = (filter.modeInput?.value || "").trim();
          const normalizedFilterValues = getFilterValues(filter)
            .map((value) => value.toLowerCase());
          if (!mode) {{
            return true;
          }}

          const cellValue = (cells[columnIndex]?.textContent || "").trim();
          const normalizedCellValue = cellValue.toLowerCase();

          if (mode === "duplicates") {{
            const counts = duplicateMaps.get(columnName);
            return Boolean(cellValue) && (counts?.get(cellValue) || 0) > 1;
          }}
          if (!normalizedFilterValues.length) {{
            return true;
          }}
          if (mode === "contains") {{
            return normalizedFilterValues.some((value) => normalizedCellValue.includes(value));
          }}
          if (mode === "not_contains") {{
            return normalizedFilterValues.every((value) => !normalizedCellValue.includes(value));
          }}

          return normalizedFilterValues.includes(normalizedCellValue);
        }});

        const match = queryMatch && filterMatch;
        row.style.display = match ? "" : "none";
      }});

      const visibleRows = getVisibleRows();
      updateSummary(visibleRows);
      if (exportButton) {{
        exportButton.disabled = visibleRows.length === 0;
      }}
      writeFilterStateToForm();

      const activeFilters = filterPairs
        .map((filter) => {{
          const mode = (filter.modeInput?.value || "").trim();
          if (!mode) {{
            return "";
          }}
          const label = filter.columnName.replace(/_/g, " ");
          if (mode === "duplicates") {{
            return label + ": duplicates";
          }}
          const terms = getFilterValues(filter);
          if (!terms.length) {{
            return "";
          }}
          const modeLabel = mode === "not_contains" ? "does not contain" : mode;
          return label + ": " + modeLabel + " " + terms.join(" or ");
        }})
        .filter(Boolean);

      const parts = [];
      parts.push("Showing " + visibleRows.length + " of " + rows.length + " row(s)");
      if (query) {{
        parts.push('search: "' + input.value.trim() + '"');
      }}
      if (activeFilters.length) {{
        parts.push(activeFilters.join(" | "));
      }}
      meta.textContent = parts.join(" | ");
    }};

    input.addEventListener("input", update);
    filterPairs.forEach((filter) => {{
      filter.modeInput?.addEventListener("change", update);
    }});
    addFilterButtons.forEach((button) => {{
      button.addEventListener("click", () => {{
        const columnName = (button.dataset.column || "").trim().toLowerCase();
        const filter = filterPairs.find((item) => item.columnName === columnName);
        const datalistId = "filter-" + columnName + "-options";
        if (!filter?.valuesContainer) {{
          return;
        }}
        const currentCount = filter.valuesContainer.querySelectorAll('[data-role="filter-term"]').length;
        const inputNode = document.createElement("input");
        inputNode.type = "text";
        inputNode.setAttribute("data-column", columnName);
        inputNode.setAttribute("data-role", "filter-term");
        inputNode.setAttribute("list", datalistId);
        inputNode.placeholder = "Value " + String(currentCount + 1) + " (optional)";
        inputNode.addEventListener("input", update);
        filter.valuesContainer.appendChild(inputNode);
        inputNode.focus();
      }});
    }});
    filterPairs.forEach((filter) => {{
      Array.from(filter.valuesContainer?.querySelectorAll('[data-role="filter-term"]') || []).forEach((inputNode) => {{
        inputNode.addEventListener("input", update);
      }});
    }});
    exportButton?.addEventListener("click", exportFilteredRows);
    addFilteredPrintButton?.addEventListener("click", () => {{
      writeFilterStateToForm();
      filteredPrintForm?.submit();
    }});
    rows.forEach((row) => {{
      const checkbox = row.querySelector(".row-check");
      if (!checkbox) {{
        return;
      }}
      checkbox.addEventListener("change", () => {{
        row.classList.toggle("row-selected", checkbox.checked);
      }});
    }});
    update();
  }})();
</script>
</body>
</html>"""


class ReportHandler(BaseHTTPRequestHandler):
    def send_bytes_response(
        self,
        status: int,
        content: bytes,
        content_type: str,
        content_disposition: str | None = None,
    ) -> None:
        try:
            self.send_response(status)
            self.send_header("Content-Type", content_type)
            if content_disposition:
                self.send_header("Content-Disposition", content_disposition)
            self.send_header("Content-Length", str(len(content)))
            self.end_headers()
            self.wfile.write(content)
        except (BrokenPipeError, ConnectionAbortedError, ConnectionResetError):
            return

    def send_redirect(self, location: str, status: int = 303) -> None:
        try:
            self.send_response(status)
            self.send_header("Location", location)
            self.send_header("Content-Length", "0")
            self.end_headers()
        except (BrokenPipeError, ConnectionAbortedError, ConnectionResetError):
            return

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        query = parse_qs(parsed.query)
        if parsed.path == "/download":
            self.handle_download(parsed)
            return
        if parsed.path == "/download-print-report":
            try:
                self.handle_download_print_report(parsed)
            except Exception as exc:
                page = render_print_report(error=str(exc)).encode("utf-8")
                self.send_bytes_response(400, page, "text/html; charset=utf-8")
            return
        if parsed.path == "/export-filtered":
            self.handle_export_filtered(parsed)
            return
        if parsed.path == "/print-logo":
            logo_path = get_print_logo_path()
            if logo_path is None:
                self.send_response(404)
                self.end_headers()
                return
            content_types = {
                ".png": "image/png",
                ".jpg": "image/jpeg",
                ".jpeg": "image/jpeg",
                ".svg": "image/svg+xml",
                ".webp": "image/webp",
            }
            content = logo_path.read_bytes()
            self.send_bytes_response(
                200,
                content,
                content_types.get(logo_path.suffix.lower(), "application/octet-stream"),
            )
            return
        if parsed.path == "/add-print-report":
            report_dir = get_active_report_dir()
            entry = add_current_report_to_print_queue(report_dir)
            report_name = query.get("report", [REPORT_FILES[0]])[0]
            message = (
                f"Saved branch {entry['branch']} to the printable report. "
                f"Inactive: {entry['inactive_to_update']}, "
                f"Active: {entry['active_to_update']}, "
                f"New active: {entry['new_active_employees']}, "
                f"New in system: {entry['new_in_system_since_last_month']}."
            )
            target = (
                f"/?tab=compare-employees&report={quote(report_name)}"
                f"&message={quote(message)}"
            )
            self.send_redirect(target)
            return
        if parsed.path == "/print-report-editor":
            page = render_print_report_editor(message=query.get("message", [""])[0], error=query.get("error", [""])[0]).encode("utf-8")
            self.send_bytes_response(200, page, "text/html; charset=utf-8")
            return
        if parsed.path == "/print-report":
            auto_print = query.get("autoprint", ["0"])[0].strip() in {"1", "true", "yes"}
            page = render_print_report(auto_print=auto_print).encode("utf-8")
            self.send_bytes_response(200, page, "text/html; charset=utf-8")
            return

        selected_report = query.get("report", [REPORT_FILES[0]])[0]
        active_tab = query.get("tab", ["compare-employees"])[0]
        message = query.get("message", [""])[0]
        error = query.get("error", [""])[0]

        page = render_page(active_tab, selected_report, message=message, error=error).encode("utf-8")
        self.send_bytes_response(200, page, "text/html; charset=utf-8")

    def handle_download(self, parsed) -> None:
        report_dir = get_active_report_dir()
        query = parse_qs(parsed.query)
        requested = query.get("file", [""])[0]

        file_path = report_dir / Path(requested).name
        if not file_path.exists():
            self.send_response(404)
            self.end_headers()
            return

        content = file_path.read_bytes()
        self.send_bytes_response(
            200,
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f'attachment; filename="{file_path.name}"',
        )

    def handle_download_print_report(self, parsed) -> None:
        query = parse_qs(parsed.query)
        export_format = query.get("format", ["xlsx"])[0].strip().lower()
        if export_format == "pdf":
            content = export_print_report_pdf()
            self.send_bytes_response(
                200,
                content,
                "application/pdf",
                'attachment; filename="printable_form_result.pdf"',
            )
            return

        content = export_print_report_workbook()
        self.send_bytes_response(
            200,
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            'attachment; filename="printable_form_result.xlsx"',
        )

    def handle_export_filtered(self, parsed) -> None:
        report_dir = get_active_report_dir()
        query = parse_qs(parsed.query)
        report_name = query.get("report", [REPORT_FILES[0]])[0]
        if report_name not in REPORT_FILES:
            report_name = REPORT_FILES[0]

        df = load_report(report_name, report_dir)
        filtered = filter_report_dataframe(
            df,
            search=query.get("search", [""])[0],
            department_mode=query.get("department_mode", [""])[0],
            department_values=parse_filter_values(query.get("department_values", [""])[0]),
            position_mode=query.get("position_mode", [""])[0],
            position_values=parse_filter_values(query.get("position_values", [""])[0]),
        )
        export_format = query.get("format", ["csv"])[0].strip().lower()
        if export_format not in {"csv", "xlsx"}:
            export_format = "csv"

        content, content_type, filename = export_report_dataframe(
            filtered,
            report_name,
            export_format,
        )
        self.send_bytes_response(
            200,
            content,
            content_type,
            f'attachment; filename="{filename}"',
        )

    def do_POST(self) -> None:
        active_tab = "compare-employees"
        selected_report = REPORT_FILES[0]
        action = ""
        try:
            files = parse_multipart(self)
            action = files.get("action", ("", b""))[1].decode("utf-8", errors="ignore").strip()
            if action == "remove_print_row":
                raw_index = files.get("row_index", ("", b""))[1].decode("utf-8", errors="ignore").strip()
                removed = remove_print_report_entry(int(raw_index))
                if not removed:
                    raise ValueError("Printable report row was not found.")
                page = render_print_report(message="Removed item from printable report.").encode("utf-8")
                self.send_bytes_response(200, page, "text/html; charset=utf-8")
                return
            if action == "clear_print_queue":
                clear_print_report_queue()
                page = render_print_report(message="Cleared printable report data.").encode("utf-8")
                self.send_bytes_response(200, page, "text/html; charset=utf-8")
                return
            if action == "save_and_reset_print_queue":
                saved_path = save_current_print_report_archive()
                clear_print_report_queue()
                page = render_print_report(
                    message=(
                        f"Saved current printable form to {saved_path.name} in "
                        f"{saved_path.parent.name}. Printable form data was reset."
                    )
                ).encode("utf-8")
                self.send_bytes_response(200, page, "text/html; charset=utf-8")
                return
            if action == "upload_logo":
                if "logo_file" not in files or not files["logo_file"][0]:
                    raise ValueError("Choose a logo file to upload.")
                save_print_logo(files["logo_file"])
                page = render_print_report(message="Logo uploaded for printable report.").encode("utf-8")
                self.send_bytes_response(200, page, "text/html; charset=utf-8")
                return
            if action == "save_print_preset":
                preset = build_print_preset_from_form(files)
                state = load_print_presets()
                presets = state["presets"]
                presets[preset["preset_name"]] = preset
                save_print_presets(preset["preset_name"], presets)
                page = render_print_report_editor(message=f"Saved and applied print preset: {preset['preset_name']}.").encode("utf-8")
                self.send_bytes_response(200, page, "text/html; charset=utf-8")
                return
            if action == "apply_print_preset":
                preset_name = files.get("apply_preset_name", ("", b""))[1].decode("utf-8", errors="ignore").strip()
                state = load_print_presets()
                presets = state["presets"]
                if preset_name not in presets:
                    raise ValueError("Selected print preset was not found.")
                save_print_presets(preset_name, presets)
                page = render_print_report_editor(message=f"Applied print preset: {preset_name}.").encode("utf-8")
                self.send_bytes_response(200, page, "text/html; charset=utf-8")
                return
            if action == "upload_updated_system":
                report_name = files.get("report", ("", b""))[1].decode("utf-8", errors="ignore").strip()
                if report_name not in REPORT_FILES:
                    report_name = REPORT_FILES[0]
                if "updated_system_file" not in files or not files["updated_system_file"][0]:
                    raise ValueError("Choose an updated system_clean file to upload.")
                report_dir = get_active_report_dir()
                if report_dir == BASE_DIR:
                    raise ValueError("Run or open a branch comparison first before uploading the updated system file.")
                uploaded_name = Path(files["updated_system_file"][0]).name
                suffix = Path(uploaded_name).suffix.lower()
                if suffix not in {".xlsx", ".xls"}:
                    raise ValueError("Updated system file must be an XLS or XLSX workbook.")
                for filename in UPDATED_SYSTEM_FILENAMES:
                    existing = report_dir / filename
                    if existing.exists():
                        existing.unlink()
                save_uploaded_file("updated_system_clean", files["updated_system_file"], report_dir)
                status_override = files.get("updated_status_column", ("", b""))[1].decode("utf-8", errors="ignore").strip()
                branch_override = files.get("updated_branch_column", ("", b""))[1].decode("utf-8", errors="ignore").strip()
                save_updated_system_config(report_dir, status_override, branch_override)
                status_summary = summarize_updated_system(report_dir)
                message = (
                    f"Uploaded updated system_clean for {report_dir.name}. "
                    f"Inactive: {status_summary['inactive']}, "
                    f"Active: {status_summary['active']}, "
                    f"Overall: {status_summary['overall']}."
                )
                target = (
                    f"/?tab=compare-employees&report={quote(report_name)}"
                    f"&message={quote(message)}"
                )
                self.send_redirect(target)
                return
            if action == "load_existing_run":
                report_name = files.get("report", ("", b""))[1].decode("utf-8", errors="ignore").strip()
                if report_name not in REPORT_FILES:
                    report_name = REPORT_FILES[0]
                run_name = files.get("existing_run", ("", b""))[1].decode("utf-8", errors="ignore").strip()
                run_dir = get_saved_report_run(run_name)
                if run_dir is None:
                    raise ValueError("Choose a valid previous comparison to preview.")
                set_active_report_dir(run_dir)
                message = f"Loaded existing comparison: {run_dir.name}."
                target = (
                    f"/?tab=compare-employees&report={quote(report_name)}"
                    f"&message={quote(message)}"
                )
                self.send_redirect(target)
                return
            if action == "add_print_report":
                report_dir = get_active_report_dir()
                entry = add_current_report_to_print_queue(report_dir)
                report_name = files.get("report", ("", b""))[1].decode("utf-8", errors="ignore").strip()
                if report_name not in REPORT_FILES:
                    report_name = REPORT_FILES[0]
                message = (
                f"Saved branch {entry['branch']} to the printable report. "
                f"Inactive: {entry['inactive_to_update']}, "
                f"Active: {entry['active_to_update']}, "
                f"New active: {entry['new_active_employees']}, "
                f"New in system: {entry['new_in_system_since_last_month']}."
            )
                target = (
                    f"/?tab=compare-employees&report={quote(report_name)}"
                    f"&message={quote(message)}"
                )
                self.send_redirect(target)
                return
            if action == "add_status_summary_to_print_report":
                report_dir = get_active_report_dir()
                report_name = files.get("report", ("", b""))[1].decode("utf-8", errors="ignore").strip()
                if report_name not in REPORT_FILES:
                    report_name = REPORT_FILES[0]
                status_summary = upsert_status_summary_to_print_queue(report_dir)
                message = (
                    f"Saved updated system branch breakdown to the printable report. "
                    f"Branches: {len(status_summary['breakdown'])}, "
                    f"Inactive: {status_summary['inactive']}, "
                    f"Active: {status_summary['active']}, "
                    f"Overall: {status_summary['overall']}."
                )
                target = (
                    f"/?tab=compare-employees&report={quote(report_name)}"
                    f"&message={quote(message)}"
                )
                self.send_redirect(target)
                return
            if action == "add_filtered_print_report":
                report_dir = get_active_report_dir()
                report_name = files.get("report", ("", b""))[1].decode("utf-8", errors="ignore").strip()
                if report_name not in REPORT_FILES:
                    report_name = REPORT_FILES[0]
                df = load_report(report_name, report_dir)
                filtered = filter_report_dataframe(
                    df,
                    search=files.get("search", ("", b""))[1].decode("utf-8", errors="ignore").strip(),
                    department_mode=files.get("department_mode", ("", b""))[1].decode("utf-8", errors="ignore").strip(),
                    department_values=parse_filter_values(
                        files.get("department_values", ("", b""))[1].decode("utf-8", errors="ignore")
                    ),
                    position_mode=files.get("position_mode", ("", b""))[1].decode("utf-8", errors="ignore").strip(),
                    position_values=parse_filter_values(
                        files.get("position_values", ("", b""))[1].decode("utf-8", errors="ignore")
                    ),
                )
                entry = upsert_filtered_report_to_print_queue(report_dir, report_name, len(filtered.index))
                report_label = REPORT_LABELS.get(report_name, report_name)
                message = (
                    f"Saved filtered {report_label} count for {entry['branch']} to the printable report. "
                    f"Filtered rows: {len(filtered.index)}."
                )
                target = (
                    f"/?tab=compare-employees&report={quote(report_name)}"
                    f"&message={quote(message)}"
                )
                self.send_redirect(target)
                return

            if "system_file" not in files or "hr_file" not in files:
                raise ValueError("Upload both the system file and the HR file.")

            run_dir = UPLOAD_DIR / f"run_{format_local_timestamp(local_now(), '%Y%m%d_%H%M%S')}"
            run_dir.mkdir(parents=True, exist_ok=True)

            system_path = save_uploaded_file("system_clean", files["system_file"], run_dir)
            hr_path = save_uploaded_file("hr_clean", files["hr_file"], run_dir)
            result = process_reports(system_path=system_path, hr_path=hr_path, output_dir=run_dir)
            set_active_report_dir(run_dir)

            message = (
                f"Compared {result['system_file']} and {result['hr_file']}. "
                f"Run branch: {result['run_branch_key']}. "
                f"Matched: {result['matched_count']}, "
                f"Inactive updates: {result['inactive_count']}, "
                f"Active updates: {result['active_count']}, "
                f"New active: {result['new_active_count']}, "
                f"New in system: {result['new_in_system_count']} "
                f"(active {result['new_in_system_active_count']}, inactive {result['new_in_system_inactive_count']}). "
                f"{'Used previous snapshot ' + result['new_in_system_previous_snapshot_name'] + '.' if result['new_in_system_previous_snapshot_found'] else 'No previous snapshot was found for this run branch, so new-in-system counting was skipped.'}"
                f"{(' Missing previous snapshot for: ' + ', '.join(result['new_in_system_missing_snapshot_branches']) + '.') if result['new_in_system_missing_snapshot_branches'] else ''}"
            )
            page = render_page(active_tab, selected_report, message=message).encode("utf-8")
        except Exception as exc:
            if action in {"remove_print_row", "clear_print_queue", "save_and_reset_print_queue", "upload_logo"}:
                page = render_print_report(error=str(exc)).encode("utf-8")
            elif action in {"save_print_preset", "apply_print_preset"}:
                page = render_print_report_editor(error=str(exc)).encode("utf-8")
            else:
                page = render_page(active_tab, selected_report, error=str(exc)).encode("utf-8")
            self.send_bytes_response(400, page, "text/html; charset=utf-8")
            return

        self.send_bytes_response(200, page, "text/html; charset=utf-8")

    def log_message(self, format: str, *args) -> None:
        return


if __name__ == "__main__":
    host = "127.0.0.1"
    port = 8000
    server = HTTPServer((host, port), ReportHandler)
    print(f"Preview server running at http://{host}:{port}")
    server.serve_forever()
